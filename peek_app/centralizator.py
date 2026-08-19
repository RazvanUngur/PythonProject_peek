# =============================================================================
# centralizator.py — Citire și actualizare 0_Centralizator_PEEK-VEK.xlsx
# =============================================================================
# Exportă:
#   update_centralizator(excel_path, site_id, central_folder)
#       — actualizează un SINGUR contor (compatibilitate; rescrie fișierul).
#   update_centralizator_batch(rezultate, central_folder, log_callback=None)
#       — actualizează TOATE contoarele dintr-o rulare cu O SINGURĂ rescriere
#         a fișierului. De preferat la procesare în masă (vezi app.py) —
#         apelul update_centralizator() în buclă rescrie integral fișierul
#         (inclusiv sheet-urile Contoare/Analiza/Reguli Calcul + backup) de
#         N ori, ceea ce devine foarte costisitor pe măsură ce fișierul
#         crește (comportament O(N²)).
# =============================================================================

import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    CENTRAL_FILE_NAME, CENTRAL_FILE_FOLDER,
    VEHICLE_ANALYSIS, MIN_LUNI_AN, MIN_LUNI_AN_MAI,
    MIN_ORE_ZI, CONTOARE_HEADERS,
)
from database import get_contoare_db

# ══════════════════════════════════════════════════════════════════════════════
# PASUL 1 — această funcție trebuie să fie ÎNAINTE de update_centralizator
# ══════════════════════════════════════════════════════════════════════════════
def _read_lunar_sheet_from_report(excel_path, site_id):
    """
    Citește corect sheet-ul 'Media Zilnica Lunara' dintr-un raport individual.
    """
    luna_map = {'Ian':1,'Feb':2,'Mar':3,'Apr':4,'Mai':5,'Iun':6,
                'Iul':7,'Aug':8,'Sep':9,'Oct':10,'Noi':11,'Dec':12}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Media Zilnica Lunara"]

    header_row = None
    col_map    = {}

    for r in range(1, ws.max_row + 1):
        row_data = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is not None:
                key = str(val).strip()
                if key not in row_data:
                    row_data[key] = c

        an_col_found   = row_data.get("An",   999)
        luna_col_found = row_data.get("Luna", 999)
        if an_col_found <= 10 and luna_col_found <= 10:
            header_row = r
            col_map    = row_data
            break

    if header_row is None:
        wb.close()
        raise RuntimeError(
            f"Nu am găsit headerul 'An'+'Luna' în 'Media Zilnica Lunara' "
            f"al fișierului {excel_path}")

    def to_an(v):
        if v is None: return None
        try:
            n = int(float(str(v)))
            return n if 2000 <= n <= 2100 else None
        except Exception: return None

    def to_luna(v):
        if v is None: return None
        s = str(v).strip()
        if s in luna_map: return luna_map[s]
        try:
            n = int(float(s))
            return n if 1 <= n <= 12 else None
        except Exception: return None

    wanted = ['Post', 'An', 'Luna',
              'Clasa 1','Clasa 2','Clasa 3','Clasa 4','Clasa 5',
              'Clasa 6','Clasa 7','Clasa 8','Clasa 15',
              'Total', 'Indicator', 'Mod de funcționare', 'Zile cu înregistrări']
    present_cols = {k: v for k, v in col_map.items() if k in wanted}

    rows_out = []
    for r in range(header_row + 1, ws.max_row + 1):
        an_col   = present_cols.get('An')
        luna_col = present_cols.get('Luna')
        if not an_col or not luna_col:
            continue

        an_val   = to_an(ws.cell(r, an_col).value)
        luna_val = to_luna(ws.cell(r, luna_col).value)

        if an_val is None or luna_val is None:
            continue

        row_dict = {'Contor': str(site_id), 'An': an_val, 'Luna': luna_val}
        for col_name, col_idx in present_cols.items():
            if col_name not in ('An', 'Luna'):
                row_dict[col_name] = ws.cell(r, col_idx).value
        rows_out.append(row_dict)

    wb.close()

    if not rows_out:
        raise RuntimeError(
            f"Nu am găsit rânduri valide în 'Media Zilnica Lunara' "
            f"al fișierului {excel_path}")

    df = pd.DataFrame(rows_out)
    df = df.rename(columns={c: c.replace('Clasa ', 'Clasa_')
                             for c in df.columns if c.startswith('Clasa ')})
    if 'Post' in df.columns:
        df = df.drop(columns=['Post'])

    for group_name, clase in VEHICLE_ANALYSIS.items():
        if group_name not in df.columns:
            cols_found = [c for c in clase if c in df.columns]
            df[group_name] = df[cols_found].sum(axis=1) if cols_found else 0

    clase_cols     = [c for c in ['Clasa_1','Clasa_2','Clasa_3','Clasa_4','Clasa_5',
                                   'Clasa_6','Clasa_7','Clasa_8','Clasa_15']
                      if c in df.columns]
    veh_cols       = [g for g in VEHICLE_ANALYSIS.keys() if g in df.columns]
    total_col      = ['Total'] if 'Total' in df.columns else []
    indicator_cols = [c for c in ['Indicator', 'Mod de funcționare',
                                   'Zile cu înregistrări'] if c in df.columns]

    keep = ['Contor', 'An', 'Luna'] + clase_cols + total_col + veh_cols + indicator_cols
    keep = [c for c in keep if c in df.columns]
    return df[keep].copy()


# ══════════════════════════════════════════════════════════════════════════════
# PASUL 2 — update_centralizator (cu sheet "Contoare" persistent de la col B)
# ══════════════════════════════════════════════════════════════════════════════
def _write_centralizator_file(df_final, _last_user):
    """
    Scrie/rescrie fișierul 0_Centralizator_PEEK-VEK.xlsx complet (sheet-urile
    "Media Zilnica Lunara", "Contoare", "Prelucrare manuala", "Analiza" cu
    grafice, "Reguli Calcul") pornind de la df_final — datele deja consolidate
    pentru TOATE contoarele. Face și backup-ul zilnic în "Istoric Procesari".

    Acesta e pasul costisitor (rescriere completă + grafice + backup), motiv
    pentru care trebuie apelat O SINGURĂ DATĂ per rulare, nu per contor —
    vezi update_centralizator_batch().
    """
    C_DARK   = "1F4E79"
    C_MID    = "2E75B6"
    C_LIGHT  = "D6E4F0"
    C_WHITE  = "FFFFFF"
    C_YELLOW = "FFF2CC"
    C_ORANGE = "FCE4D6"
    C_VIOLET = "EAD1DC"   # violet pal — Prelucrare manuală MZL

    def fill(hex_c):
        return PatternFill('solid', start_color=hex_c)

    thin = Side(style='thin', color='BFBFBF')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfont(sz=10):
        return Font(name='Arial', size=sz, bold=True, color=C_WHITE)

    def dfont(sz=10, bold=False, color="1F1F1F"):
        return Font(name='Arial', size=sz, bold=bold, color=color)

    def ctr():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def rgt():
        return Alignment(horizontal='right', vertical='center')

    # Centralizatorul e mereu la calea fixă, indiferent de unde vin fișierele .bin
    os.makedirs(CENTRAL_FILE_FOLDER, exist_ok=True)
    CENTRAL_FILE = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)

    # ── 4. Citim baza de date Contoare din SQLite (sursa unică de adevăr) ──────
    # contoare.db este sursa principală — nu mai depindem de snapshot-ul Excel.
    try:
        db_contoare_snapshot = get_contoare_db().get_all()
    except Exception as _e_cdb:
        print(f"[WARN] Nu am putut citi contoare.db: {_e_cdb}. Sheet 'Contoare' va fi gol.")
        db_contoare_snapshot = {}

    # ── 5. Scriere Excel ──────────────────────────────────────────────────────
    with pd.ExcelWriter(CENTRAL_FILE, engine="openpyxl") as writer:
        df_final.to_excel(writer, sheet_name="Media Zilnica Lunara", index=False)
        wb = writer.book
        ws = writer.sheets["Media Zilnica Lunara"]

        # Titlu rândul 1
        ws.insert_rows(1)
        n_cols = max(len(df_final.columns), 10)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        ws["A1"] = "CENTRALIZATOR PEEK  |  Media Zilnică Lunară pe Contoare"
        ws["A1"].font      = Font(name='Arial', size=13, bold=True, color=C_WHITE)
        ws["A1"].fill      = fill(C_DARK)
        ws["A1"].alignment = ctr()
        ws.row_dimensions[1].height = 28

        # Header rândul 2
        for cell in ws[2]:
            if cell.value:
                cell.font      = hfont(9)
                cell.fill      = fill(C_MID)
                cell.alignment = ctr()
                cell.border    = brd
        ws.row_dimensions[2].height = 28

        contor_col_idx = None
        for cell in ws[2]:
            if cell.value and str(cell.value).strip() == 'Contor':
                contor_col_idx = cell.column
                break

        indicator_col_idx = None
        mod_func_col_idx  = None
        for cell in ws[2]:
            if cell.value and str(cell.value).strip() == 'Indicator':
                indicator_col_idx = cell.column
            if cell.value and str(cell.value).strip() == 'Mod de funcționare':
                mod_func_col_idx = cell.column

        contoare_unice = list(dict.fromkeys(
            ws.cell(r, contor_col_idx).value
            for r in range(3, ws.max_row + 1)
            if contor_col_idx and ws.cell(r, contor_col_idx).value
        ))
        contoare_lst = sorted(df_final["Contor"].astype(str).unique().tolist())
        paleta = [C_LIGHT, "E8F4FD", "EBF5EB", "FDF6E3", "F5EEF8"]
        contor_colors = {str(ct): paleta[i % len(paleta)]
                         for i, ct in enumerate(contoare_unice)}

        num_headers = {'Clasa_1','Clasa_2','Clasa_3','Clasa_4','Clasa_5',
                       'Clasa_6','Clasa_7','Clasa_8','Clasa_15','Total',
                       'Autoturisme','LGV','HGV','Autobuze'}
        num_cols = set()
        for cell in ws[2]:
            if cell.value and str(cell.value).strip() in num_headers:
                num_cols.add(cell.column)

        for r in range(3, ws.max_row + 1):
            ct_val   = str(ws.cell(r, contor_col_idx).value) if contor_col_idx else ""
            row_fill = fill(contor_colors.get(ct_val, C_LIGHT))

            if indicator_col_idx:
                ind_val = str(ws.cell(r, indicator_col_idx).value or "")
                if "manuală" in ind_val or "manuala" in ind_val.lower():
                    row_fill = fill(C_VIOLET)
                elif "parțiale" in ind_val:
                    row_fill = fill(C_YELLOW)
                elif "Nu există" in ind_val:
                    row_fill = fill(C_ORANGE)

            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                cell.font      = dfont(9)
                cell.fill      = row_fill
                cell.alignment = ctr()
                cell.border    = brd
                if c in num_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                # Stilizare specială coloana Mod de funcționare
                if mod_func_col_idx and c == mod_func_col_idx:
                    _mf = str(cell.value or "").lower()
                    if "clasificator" in _mf:
                        cell.font = Font(name='Arial', size=9, bold=True,
                                         color="1E8449")
                    elif "totalizator" in _mf:
                        cell.font = Font(name='Arial', size=9, bold=True,
                                         color="7D6608")
                    elif "null" in _mf or "neutiliz" in _mf:
                        cell.font = Font(name='Arial', size=9, bold=True,
                                         color="922B21")

        for col_cells in ws.iter_cols(min_row=2, max_row=ws.max_row):
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=8
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 3, 30)

        ws.freeze_panes = "A3"

        # ════════════════════════════════════════════════════════════════════
        # SHEET "Contoare" — 9 coloane: Contor, Drum, Poziție km, Localitate,
        #                     Tip, IP, Lat, Lng, DRDP  (sursa: contoare.db)
        # ════════════════════════════════════════════════════════════════════
        ws_contoare = wb.create_sheet("Contoare")
        ws_contoare.sheet_view.showGridLines = False

        # contoare.db este sursa unică de adevăr — folosim direct ce am citit din SQLite.
        db_contoare = db_contoare_snapshot

        # Contoare noi (apar în raport dar lipsesc din contoare.db) → upsert în SQLite
        _cdb = None
        try:
            _cdb = get_contoare_db()
        except Exception:
            pass
        for ct in contoare_lst:
            if ct not in db_contoare:
                _empty = {"Drum": "", "Pozitie_km": "", "Localitate": "",
                          "Tip": "", "IP": "", "lat": None, "lng": None, "DRDP": ""}
                db_contoare[ct] = _empty
                if _cdb is not None:
                    try:
                        _cdb.upsert(ct, _empty)
                    except Exception as _eu:
                        print(f"[WARN] Nu am putut adăuga contorul {ct} în contoare.db: {_eu}")

        _CT_HEADERS    = ["Contor", "Drum", "Poziție km", "Localitate", "Tip", "IP",
                           "Lat", "Lng", "DRDP"]
        _CT_WIDTHS     = [16, 22, 14, 22, 26, 18, 14, 14, 16]
        for c_idx, hdr in enumerate(_CT_HEADERS, 1):
            cell = ws_contoare.cell(1, c_idx, hdr)
            cell.font      = hfont(10)
            cell.fill      = fill(C_DARK if c_idx == 1 else C_MID)
            cell.alignment = ctr()
            cell.border    = brd
            ws_contoare.column_dimensions[get_column_letter(c_idx)].width = _CT_WIDTHS[c_idx - 1]
        ws_contoare.row_dimensions[1].height = 26

        for r_idx, (ct_id, data) in enumerate(sorted(db_contoare.items()), 2):
            row_fill = fill(paleta[(r_idx - 2) % len(paleta)])
            lat_val = data.get("lat") or data.get("Lat")
            lng_val = data.get("lng") or data.get("Lng")
            values = [
                ct_id,
                data.get("Drum", ""),
                data.get("Pozitie_km", ""),
                data.get("Localitate", ""),
                data.get("Tip", ""),
                data.get("IP", ""),
                round(float(lat_val), 6) if lat_val not in (None, "") else "",
                round(float(lng_val), 6) if lng_val not in (None, "") else "",
                data.get("DRDP", ""),
            ]
            for c_idx, val in enumerate(values, 1):
                cell = ws_contoare.cell(r_idx, c_idx, val)
                cell.font      = dfont(10, bold=(c_idx == 1))
                cell.fill      = row_fill
                cell.alignment = ctr()
                cell.border    = brd

        ws_contoare.freeze_panes = "A2"

        # ════════════════════════════════════════════════════════════════════
        # SHEET „Prelucrare manuala" — istoric complet din SQLite (read-only)
        # ════════════════════════════════════════════════════════════════════
        # Reconstruit complet la fiecare actualizare din tabelul mzl_manual.
        # Editarea se face exclusiv din GUI (buton „Procesare manuală MZL").
        # ════════════════════════════════════════════════════════════════════

        if "Prelucrare manuala" in wb.sheetnames:
            del wb["Prelucrare manuala"]

        ws_pm = wb.create_sheet("Prelucrare manuala")
        ws_pm.sheet_view.showGridLines = False

        LUNI_NUME_PM = {
            1: "Ianuarie", 2: "Februarie", 3: "Martie",    4: "Aprilie",
            5: "Mai",      6: "Iunie",     7: "Iulie",      8: "August",
            9: "Septembrie", 10: "Octombrie", 11: "Noiembrie", 12: "Decembrie",
        }

        # ── Citire date din SQLite ────────────────────────────────────────────
        df_mzl = pd.DataFrame()
        _db_ok = False
        try:
            from database import get_traffic_db
            _tdb = get_traffic_db()
            df_mzl = _tdb.get_all_mzl_manual()
            _db_ok = True
        except Exception as _e_db:
            print(f"[WARN] SQLite mzl_manual indisponibil: {_e_db}")

        _now_pm = __import__('datetime').datetime.now().strftime("%d.%m.%Y %H:%M")
        _n_cols_pm = 8

        # ── Titlu ─────────────────────────────────────────────────────────────
        ws_pm.merge_cells(f"A1:{get_column_letter(_n_cols_pm)}1")
        _t = ws_pm["A1"]
        _t.value     = "PRELUCRARE MANUALĂ MZL  |  Istoric modificări operatori"
        _t.font      = Font(name='Arial', size=13, bold=True, color=C_WHITE)
        _t.fill      = fill(C_DARK)
        _t.alignment = ctr()
        ws_pm.row_dimensions[1].height = 28

        # ── Subtitlu ──────────────────────────────────────────────────────────
        ws_pm.merge_cells(f"A2:{get_column_letter(_n_cols_pm)}2")
        _sub = ws_pm["A2"]
        if _db_ok:
            _sub.value = (
                f"Generat automat din SQLite (trafic.db → mzl_manual)  |  "
                f"{len(df_mzl)} modificări înregistrate  |  "
                f"Actualizat: {_now_pm}  |  Utilizator: {_last_user}"
            )
        else:
            _sub.value = (
                "⚠ SQLite indisponibil — instalați database.py. "
                "Editare din butonul 'Procesare manuală MZL' din GUI."
            )
        _sub.font      = Font(name='Arial', size=9, italic=True, color="595959")
        _sub.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        _sub.fill      = fill("F2F2F2")
        ws_pm.row_dimensions[2].height = 22

        # ── Header tabel ──────────────────────────────────────────────────────
        _pm_hdrs = ["Contor", "An", "Lună",
                    "Valoare veche (auto)", "Valoare nouă (manual)",
                    "Observații", "Operator", "Modificat la"]
        _pm_wds  = [12, 8, 14, 20, 22, 36, 20, 18]
        for _ci, (_h, _w) in enumerate(zip(_pm_hdrs, _pm_wds), 1):
            _hc = ws_pm.cell(3, _ci, _h)
            _hc.font      = hfont(10)
            _hc.fill      = fill("5B2C6F")   # violet închis — distincție vizuală
            _hc.alignment = ctr()
            _hc.border    = brd
            ws_pm.column_dimensions[get_column_letter(_ci)].width = _w
        ws_pm.row_dimensions[3].height = 26

        # ── Rânduri date ──────────────────────────────────────────────────────
        _r_pm = 4
        if _db_ok and not df_mzl.empty:
            # Calculăm valoarea veche (MZL auto) pentru fiecare intrare
            import calendar as _cal
            try:
                _conn_r = _tdb._conn()
                _auto_dict = {}
                for _, _rm in df_mzl.iterrows():
                    # MZL auto = media ZILNICĂ (suma totală / zile valide ≥ MIN_ORE_ZI)
                    # identic cu logica din excel_report.py
                    _rows_zi = _conn_r.execute("""
                        SELECT zi,
                               SUM(total_general) AS total_zi,
                               COUNT(*) AS ore_zi
                        FROM inregistrari_orare
                        WHERE contor=? AND an=? AND luna=?
                        GROUP BY zi
                    """, (str(_rm["contor"]), int(_rm["an"]),
                          int(_rm["luna"]))).fetchall()
                    _zile_val = [r for r in _rows_zi if int(r["ore_zi"] or 0) >= 22]
                    _key = (str(_rm["contor"]), int(_rm["an"]), int(_rm["luna"]))
                    if _zile_val:
                        _tot_val = sum(int(r["total_zi"] or 0) for r in _zile_val)
                        _mzl_calc = round(_tot_val / len(_zile_val))
                        _auto_dict[_key] = {
                            "mzl_auto": _mzl_calc,
                            "zile": len(_zile_val)
                        }
                    else:
                        _auto_dict[_key] = None
            except Exception:
                _auto_dict = {}

            _df_sort = df_mzl.sort_values(
                ["contor", "an", "luna"],
                ascending=[True, False, False]
            ).reset_index(drop=True)

            for _idx, _rm in _df_sort.iterrows():
                _ct   = str(_rm["contor"])
                _an   = int(_rm["an"])
                _lu   = int(_rm["luna"])
                _mnew = _rm["mzl_valoare"]
                _obs  = str(_rm.get("observatii", "") or "")
                _op   = str(_rm.get("utilizator", "") or "")
                _mla  = str(_rm.get("modificat_la", "") or "")

                # Format data
                try:
                    from datetime import datetime as _dtt
                    _mla_fmt = _dtt.fromisoformat(_mla).strftime("%d.%m.%Y %H:%M")
                except Exception:
                    _mla_fmt = _mla

                # Valoare veche
                _av = _auto_dict.get((_ct, _an, _lu))
                if _av and _av["mzl_auto"] is not None:
                    _zile_luna = _cal.monthrange(_an, _lu)[1]
                    _val_v = int(_av["mzl_auto"])
                    _val_v_str = f"{_val_v:,}  ({int(_av['zile'])}/{_zile_luna} zile)"
                else:
                    _val_v_str = "Fără date"

                _luna_str = LUNI_NUME_PM.get(_lu, str(_lu))
                _rf_pm = fill(C_VIOLET) if _idx % 2 == 0 else fill("F5EDF8")

                _vals_pm = [
                    _ct, _an, _luna_str,
                    _val_v_str,
                    int(_mnew) if _mnew is not None else "",
                    _obs, _op, _mla_fmt,
                ]
                _aligns_pm = [
                    ctr(), ctr(), ctr(),
                    Alignment(horizontal='right',  vertical='center'),
                    Alignment(horizontal='right',  vertical='center'),
                    Alignment(horizontal='left',   vertical='center', wrap_text=True),
                    Alignment(horizontal='center', vertical='center'),
                    Alignment(horizontal='center', vertical='center'),
                ]
                for _ci2, (_v, _al) in enumerate(zip(_vals_pm, _aligns_pm), 1):
                    _cell = ws_pm.cell(_r_pm, _ci2, _v)
                    _cell.font      = dfont(9, bold=(_ci2 in (1, 5)))
                    _cell.fill      = _rf_pm
                    _cell.border    = brd
                    _cell.alignment = _al
                    if _ci2 == 5 and isinstance(_v, (int, float)):
                        _cell.number_format = "#,##0"
                    if _ci2 == 5:
                        _cell.font = Font(name='Arial', size=10, bold=True,
                                          color="4A235A")
                ws_pm.row_dimensions[_r_pm].height = 20
                _r_pm += 1
        else:
            ws_pm.merge_cells(f"A4:{get_column_letter(_n_cols_pm)}4")
            _ec = ws_pm["A4"]
            _ec.value = (
                "Nu există modificări manuale înregistrate." if _db_ok
                else "SQLite indisponibil — verificați că database.py este instalat."
            )
            _ec.font      = Font(name='Arial', size=9, italic=True, color="888888")
            _ec.alignment = Alignment(horizontal='center', vertical='center')
            _ec.fill      = fill("FAFAFA")
            ws_pm.row_dimensions[4].height = 28
            _r_pm = 5

        # ── Notă subsol ───────────────────────────────────────────────────────
        ws_pm.merge_cells(f"A{_r_pm}:{get_column_letter(_n_cols_pm)}{_r_pm}")
        _nota = ws_pm.cell(_r_pm, 1,
            "ℹ Sheet generat automat. Nu editați direct — "
            "folosiți butonul „Procesare manuală MZL” din aplicație.")
        _nota.font      = Font(name='Arial', size=8, italic=True, color="AAAAAA")
        _nota.alignment = Alignment(horizontal='left', vertical='center')

        ws_pm.freeze_panes = "A4"

        # ════════════════════════════════════════════════════════════════════
        # SHEET "Analiza" — dropdown legat dinamic de sheet-ul "Contoare"
        # ════════════════════════════════════════════════════════════════════
        ws_a = wb.create_sheet("Analiza")
        ws_a.sheet_view.showGridLines = False

        ani       = sorted(df_final["An"].dropna().unique().astype(int).tolist())
        luni_list = ["Ianuarie","Februarie","Martie","Aprilie","Mai","Iunie",
                     "Iulie","August","Septembrie","Octombrie","Noiembrie","Decembrie"]

        # ── Dropdown contor — compatibil complet cu openpyxl ─────────────────
        # Referința la alt sheet în DataValidation necesită tipul "list" cu
        # sqref setat explicit; fără showErrorMessage=False apar warnings
        dv = DataValidation(
            type="list",
            formula1="Contoare!$A$2:$A$500",  # ← FĂRĂ semnul = în față
            allow_blank=False,
            showDropDown=False,
            showErrorMessage=False,  # ← elimină warning-ul la re-citire
            showInputMessage=False,
        )
        ws_a.add_data_validation(dv)
        dv.sqref = "A1"  # ← setăm sqref direct, nu via .add()
        ws_a["A1"].value = contoare_lst[0] if contoare_lst else ""
        dv.add(ws_a["A1"])
        ws_a["A1"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                                      fill_type="solid")
        ws_a["A1"].font = Font(name='Arial', size=10, bold=True)
        ws_a["A1"].alignment = ctr()
        ws_a["A1"].border = brd
        ws_a.column_dimensions["A"].width = 14

        # B1, C1, D1 — preiau automat info din Contoare via VLOOKUP
        # (se actualizează live când utilizatorul schimbă dropdown-ul)
        vlookup_labels = {
            "B1": (2, "Drum"),
            "C1": (3, "Poziție km"),
            "D1": (4, "Localitate"),
        }
        for cell_addr, (col_idx, fallback_label) in vlookup_labels.items():
            formula = f'=IFERROR(VLOOKUP($A$1,Contoare!$A:$Z,{col_idx},0),"")'
            c = ws_a[cell_addr]
            c.value = formula
            c.font = Font(name='Arial', size=10, italic=True, color="1F4E79")
            c.fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB",
                                 fill_type="solid")
            c.alignment = ctr()
            c.border = brd
            ws_a.column_dimensions[cell_addr[0]].width = 20

        # Etichetă orientativă în E1
        ws_a["E1"].value = "← info din sheet-ul Contoare"
        ws_a["E1"].font = Font(name='Arial', size=9, color="595959", italic=True)

        ws_a.row_dimensions[1].height = 22

        col_letter_map = {}
        for cell in ws[2]:
            if cell.value:
                col_letter_map[str(cell.value).strip()] = get_column_letter(cell.column)

        COLUMN_MAP = {
            "Total vehicule": "Total",
            "Autoturisme":    "Autoturisme",
            "LGV":            "LGV",
            "HGV":            "HGV",
            "Autobuze":       "Autobuze",
        }

        culori_ani = ["2E75B6", "ED7D31", "70AD47", "FFC000", "5B9BD5"]
        start_row = 3

        indicator_letter = col_letter_map.get("Indicator", "")

        for group_name, col_name in COLUMN_MAP.items():
            col_letter = col_letter_map.get(col_name, "")
            if not col_letter:
                continue

            # ── Titlu secțiune ────────────────────────────────────────────
            tc = ws_a.cell(row=start_row - 1, column=2,
                           value=f"Media Zilnică Lunară — {group_name}")
            tc.font = Font(name='Arial', size=12, bold=True, color=C_DARK)

            # ── Header: Luna | An... ──────────────────────────────────────
            lh = ws_a.cell(row=start_row, column=2, value="Luna")
            lh.font = hfont();
            lh.fill = fill(C_MID)
            lh.border = brd;
            lh.alignment = ctr()

            for i, an in enumerate(ani):
                hc = ws_a.cell(row=start_row, column=3 + i, value=an)
                hc.font = hfont();
                hc.fill = fill(C_MID)
                hc.border = brd;
                hc.alignment = ctr()

            # ── Rânduri luni (Ianuarie–Decembrie) ────────────────────────
            for r_idx, luna in enumerate(luni_list):
                row_num = start_row + 1 + r_idx
                luna_nr = r_idx + 1
                lc = ws_a.cell(row=row_num, column=2, value=luna)
                lc.font = dfont(bold=True)
                lc.border = brd
                lc.alignment = ctr()
                lc.fill = fill(C_LIGHT) if r_idx % 2 == 0 else fill(C_WHITE)

                for c_idx, an in enumerate(ani):
                    an_cell_coord = ws_a.cell(row=start_row,
                                              column=3 + c_idx).coordinate
                    formula = (
                        f"=IFERROR(SUMIFS("
                        f"'Media Zilnica Lunara'!${col_letter}:${col_letter},"
                        f"'Media Zilnica Lunara'!$A:$A,$A$1,"
                        f"'Media Zilnica Lunara'!$B:$B,{an_cell_coord},"
                        f"'Media Zilnica Lunara'!$C:$C,{luna_nr}"
                        f"),0)"
                    )
                    fc = ws_a.cell(row=row_num, column=3 + c_idx, value=formula)
                    fc.font = dfont()
                    fc.border = brd
                    fc.alignment = rgt()
                    fc.number_format = '#,##0'
                    fc.fill = fill(C_LIGHT) if r_idx % 2 == 0 else fill(C_WHITE)

            # ── Rândul MZA — sub Decembrie ────────────────────────────────
            mza_row = start_row + 13  # rândul 1=header, 2-13=luni, 14=MZA

            lc_mza = ws_a.cell(row=mza_row, column=2, value="MZA")
            lc_mza.font = dfont(bold=True, color=C_WHITE)
            lc_mza.border = brd
            lc_mza.alignment = ctr()
            lc_mza.fill = fill(C_DARK)

            for c_idx, an in enumerate(ani):
                an_cell_coord = ws_a.cell(row=start_row,
                                          column=3 + c_idx).coordinate
                # Coordonata celulei corespunzătoare lunii Mai (luna nr.5)
                # pentru acest an și grup: rândul header + 5 (Mai = al 5-lea)
                mai_row = start_row + MIN_LUNI_AN_MAI  # +5 → rândul Mai
                mai_col = get_column_letter(3 + c_idx)
                mai_cell_coord = f"{mai_col}{mai_row}"

                if indicator_letter:
                    # Numărul de luni valide (indicator <> "Nu există date")
                    countifs_valide = (
                        f"COUNTIFS("
                        f"'Media Zilnica Lunara'!$A:$A,$A$1,"
                        f"'Media Zilnica Lunara'!$B:$B,{an_cell_coord},"
                        f"'Media Zilnica Lunara'!${indicator_letter}:${indicator_letter},"
                        f"\"<>Nu există date\""
                        f")"
                    )
                    sumifs_valide = (
                        f"SUMIFS("
                        f"'Media Zilnica Lunara'!${col_letter}:${col_letter},"
                        f"'Media Zilnica Lunara'!$A:$A,$A$1,"
                        f"'Media Zilnica Lunara'!$B:$B,{an_cell_coord},"
                        f"'Media Zilnica Lunara'!${indicator_letter}:${indicator_letter},"
                        f"\"<>Nu există date\""
                        f")"
                    )
                    # Valoarea lunii Mai pentru acest an din tabelul de mai sus
                    # (deja calculată ca SUMIFS în celula corespunzătoare)
                    formula_mza = (
                        f"=IFERROR("
                        f"IF({countifs_valide}>={MIN_LUNI_AN},"
                        # Cazul 1: ≥ MIN_LUNI_AN luni valide → media lor
                        f"{sumifs_valide}/{countifs_valide},"
                        # Cazul 2: < MIN_LUNI_AN → valoarea lunii Mai dacă există
                        f"IF({mai_cell_coord}<>0,{mai_cell_coord},\"\")"
                        f"),\"\")"
                    )
                else:
                    # Fallback fără coloana Indicator
                    mai_sumifs = (
                        f"SUMIFS("
                        f"'Media Zilnica Lunara'!${col_letter}:${col_letter},"
                        f"'Media Zilnica Lunara'!$A:$A,$A$1,"
                        f"'Media Zilnica Lunara'!$B:$B,{an_cell_coord},"
                        f"'Media Zilnica Lunara'!$C:$C,{MIN_LUNI_AN_MAI}"
                        f")"
                    )
                    formula_mza = (
                        f"=IFERROR("
                        f"IF(COUNTIFS("
                        f"'Media Zilnica Lunara'!$A:$A,$A$1,"
                        f"'Media Zilnica Lunara'!$B:$B,{an_cell_coord}"
                        f")>={MIN_LUNI_AN},"
                        f"AVERAGEIFS("
                        f"'Media Zilnica Lunara'!${col_letter}:${col_letter},"
                        f"'Media Zilnica Lunara'!$A:$A,$A$1,"
                        f"'Media Zilnica Lunara'!$B:$B,{an_cell_coord}"
                        f"),"
                        f"IF({mai_sumifs}<>0,{mai_sumifs},\"\")"
                        f"),\"\")"
                    )

                fc_mza = ws_a.cell(row=mza_row, column=3 + c_idx,
                                   value=formula_mza)
                fc_mza.font = dfont(bold=True, color=C_WHITE)
                fc_mza.border = brd
                fc_mza.alignment = rgt()
                fc_mza.number_format = '#,##0'
                fc_mza.fill = fill(C_DARK)

            # ── Lățimi coloane ────────────────────────────────────────────
            ws_a.column_dimensions["B"].width = 16
            for i in range(len(ani)):
                ws_a.column_dimensions[get_column_letter(3 + i)].width = 13

            # ── Grafic (numai lunile 1-12, MZA exclus) ───────────────────
            data_ref = Reference(ws_a, min_col=3, max_col=2 + len(ani),
                                 min_row=start_row, max_row=start_row + 12)
            cats_ref = Reference(ws_a, min_col=2,
                                 min_row=start_row + 1, max_row=start_row + 12)

            chart = LineChart()
            chart.title = f"Evoluție — {group_name}"
            chart.style = 10
            chart.width = 16
            chart.height = 9
            chart.y_axis.title = "Vehicule / zi (medie)"
            chart.x_axis.title = "Luna"
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.x_axis.delete = False
            chart.y_axis.delete = False
            chart.x_axis.axPos = "b"
            chart.x_axis.tickLblPos = "nextTo"
            chart.x_axis.textRotation = -45000
            chart.y_axis.scaling.min = 0
            chart.legend.position = 'b'
            chart.legend.overlay = False
            chart.layout = Layout(manualLayout=ManualLayout(
                x=0.08, y=0.12, h=0.72, w=0.88, xMode="edge", yMode="edge"
            ))
            for idx, serie in enumerate(chart.series):
                culoare = culori_ani[idx % len(culori_ani)]
                serie.graphicalProperties.line.solidFill = culoare
                serie.graphicalProperties.line.width = 25000
                serie.smooth = True
                serie.marker.symbol = "circle"
                serie.marker.size = 7
                serie.marker.graphicalProperties.solidFill = culoare
                serie.marker.graphicalProperties.line.solidFill = culoare

            ws_a.add_chart(chart, f"H{start_row - 1}")
            start_row += 19  # 1 titlu + 1 header + 12 luni + 1 MZA + 4 spațiu

        # ── Ordinea sheet-urilor și sheet activ ──────────────────────────
        wb.move_sheet("Contoare", offset=len(wb.sheetnames))
        for sh in wb.worksheets:
            sh.sheet_view.tabSelected = (sh.title == "Analiza")
        wb.active = wb["Analiza"]

        # ════════════════════════════════════════════════════════════════════
        # SHEET "Reguli Calcul" în Centralizator
        # ════════════════════════════════════════════════════════════════════
        ws_rc = wb.create_sheet("Reguli Calcul")
        ws_rc.sheet_view.showGridLines = False

        ws_rc.merge_cells("A1:D1")
        ws_rc["A1"] = "REGULI DE CALCUL  |  Centralizator PEEK"
        ws_rc["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
        ws_rc["A1"].fill = fill(C_DARK)
        ws_rc["A1"].alignment = ctr()
        ws_rc.row_dimensions[1].height = 28

        headers_rc = ["Regulă", "Valoare", "Unitate", "Descriere"]
        for c, h in enumerate(headers_rc, 1):
            cell = ws_rc.cell(2, c, h)
            cell.font = hfont(10)
            cell.fill = fill(C_MID)
            cell.alignment = ctr()
            cell.border = brd
        ws_rc.row_dimensions[2].height = 26

        reguli_central = [
            ("Minim ore / zi",
             22, "ore",
             "Zi validă = minim 22 ore înregistrate din 24"),
            ("Minim zile / lună",
             15, "zile",
             "Medie lunară validă = minim 15 zile valide în lună"),
            ("Alternativă săptămână",
             7, "zile",
             "Sau cel puțin 7 zile consecutive valide din lună"),
            ("Minim luni / an (MZA)",
             MIN_LUNI_AN, "luni",
             f"MZA = media aritmetică a lunilor valide, condiționat de minim {MIN_LUNI_AN} luni cu date (indicator ≠ 'Nu există date')"),
            ("Fallback MZA — luna Mai",
             MIN_LUNI_AN_MAI, "luna nr.",
             f"Dacă sunt sub {MIN_LUNI_AN} luni valide, MZA preia valoarea lunii Mai (luna {MIN_LUNI_AN_MAI}); dacă lipsește și Mai → celulă goală"),
            ("Date complete",
             "toate", "zile/lună",
             "Toate zilele lunii au date valide (ore ≥ MIN_ORE_ZI)"),
            ("Date parțiale",
             "≥15", "zile/lună",
             "Suficiente zile valide dar nu toate — se calculează media pe zilele disponibile"),
            ("Date parțiale - 7 zile",
             "1", "săptămână",
             "Cel puțin o săptămână completă (7 zile consecutive) validă — media pe 7 zile"),
            ("Nu există date (lună)",
             "0", "zile/lună",
             "Nicio zi validă în lună — luna exclusă complet din calcule MZL și MZA"),
            ("Nu există date (zi)",
             "0", "ore/zi",
             "Ziua nu are nicio oră cu vehicule înregistrate — indicator 'Nu există date', rând roșu în Rezumat Zilnic, afișat ca 0/24 la Ore înregistrate"),
        ]

        for i, (reg, val, unit, desc) in enumerate(reguli_central, 3):
            rf = fill(C_LIGHT) if i % 2 == 0 else fill(C_WHITE)
            row_data = [reg, val, unit, desc]
            for c, v in enumerate(row_data, 1):
                cell = ws_rc.cell(i, c, v)
                cell.font = dfont(9)
                cell.fill = rf
                cell.border = brd
                cell.alignment = (ctr() if c <= 3
                                  else Alignment(horizontal='left',
                                                 vertical='center',
                                                 wrap_text=True))
            ws_rc.row_dimensions[i].height = 32

        ws_rc.column_dimensions['A'].width = 28
        ws_rc.column_dimensions['B'].width = 14
        ws_rc.column_dimensions['C'].width = 12
        ws_rc.column_dimensions['D'].width = 65
        ws_rc.freeze_panes = "A3"

        # ── Ordinea finală sheet-uri ──────────────────────────────────────
        # Media Zilnica Lunara → Analiza → Contoare → Reguli Calcul
        wb.move_sheet("Contoare", offset=len(wb.sheetnames))
        wb.move_sheet("Reguli Calcul", offset=len(wb.sheetnames))
        for sh in wb.worksheets:
            sh.sheet_view.tabSelected = (sh.title == "Analiza")
        wb.active = wb["Analiza"]



    # ── Backup automat după scriere ───────────────────────────────────────────
    # Salvează o copie în: CENTRAL_FILE_FOLDER\Istoric Procesari\
    # Numele fișierului conține data în format zz-ll-aaaa, ex:
    #   0_Centralizator_PEEK_05-03-2026.xlsx
    # Se rescrie la fiecare rulare din aceeași zi.
    try:
        from datetime import datetime as _dt
        import shutil as _shutil
        _istoric_folder = os.path.join(CENTRAL_FILE_FOLDER, "Istoric Procesari")
        os.makedirs(_istoric_folder, exist_ok=True)
        _backup_date = _dt.now().strftime("%d-%m-%Y")
        _base, _ext  = os.path.splitext(CENTRAL_FILE_NAME)
        _backup_name = f"{_base}_{_backup_date}{_ext}"
        _backup_path = os.path.join(_istoric_folder, _backup_name)
        _shutil.copy2(CENTRAL_FILE, _backup_path)
    except Exception:
        pass  # backup eșuat nu oprește procesarea
    # ─────────────────────────────────────────────────────────────────────────

    return CENTRAL_FILE


# ══════════════════════════════════════════════════════════════════════════════
# PASUL 3 — update_centralizator_batch (rescriere O SINGURĂ DATĂ pentru N contoare)
# ══════════════════════════════════════════════════════════════════════════════
def update_centralizator_batch(rezultate, central_folder, log_callback=None):
    """
    Actualizează centralizatorul o SINGURĂ dată pentru toate contoarele din
    `rezultate`, în loc de o rescriere completă a fișierului per contor.

    rezultate:    listă de dict-uri cu cheile 'path' (raport Excel individual
                  al contorului) și 'id' (codul postului) — exact formatul
                  folosit deja de update_centralizator().
    central_folder: păstrat pentru compatibilitate cu semnătura veche
                  (calea reală e CENTRAL_FILE_FOLDER din config, ca și înainte).
    log_callback: opțional, func(str) — apelat cu un mesaj pentru fiecare
                  contor al cărui raport nu a putut fi citit; restul batch-ului
                  continuă normal (nu se oprește la un singur fișier stricat).

    Returnează (CENTRAL_FILE, erori), unde erori e o listă de tuple
    (site_id, exceptie) pentru contoarele care au eșuat la citire.
    """
    os.makedirs(CENTRAL_FILE_FOLDER, exist_ok=True)
    CENTRAL_FILE = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)

    # ── 0. Ultimul utilizator care a salvat fișierul (o singură citire) ────────
    _last_user = ""
    if os.path.exists(CENTRAL_FILE):
        try:
            _wb_meta = openpyxl.load_workbook(CENTRAL_FILE, read_only=True,
                                               data_only=True)
            _last_user = (_wb_meta.properties.lastModifiedBy or "").strip()
            _wb_meta.close()
        except Exception:
            _last_user = ""
    if not _last_user:
        try:
            import getpass
            _last_user = getpass.getuser()
        except Exception:
            _last_user = "Necunoscut"

    # ── 1. Date noi din TOATE rapoartele, izolat per contor ─────────────────
    # O eroare la citirea unui raport individual nu oprește restul batch-ului.
    df_new_list = []
    site_ids = set()
    erori = []
    for r in rezultate:
        try:
            df_r = _read_lunar_sheet_from_report(r["path"], r["id"])
            df_new_list.append(df_r)
            site_ids.add(str(r["id"]))
        except Exception as e_c:
            erori.append((r.get("id"), e_c))
            if log_callback:
                try:
                    log_callback(f"  ⚠ Centralizator eroare [{r.get('id')}]: {e_c}")
                except Exception:
                    pass

    if not df_new_list:
        return CENTRAL_FILE, erori

    df_new = pd.concat(df_new_list, ignore_index=True)

    # ── 2. Citire date persistente din sheet-ul "Media Zilnica Lunara" ─────────
    # La fel ca în update_centralizator(), dar excludem TOATE contoarele din
    # batch deodată (nu doar unul), pentru că vor fi rescrise cu datele noi.
    if os.path.exists(CENTRAL_FILE):
        try:
            wb_ex = openpyxl.load_workbook(CENTRAL_FILE, data_only=True)
            ws_ex = wb_ex["Media Zilnica Lunara"]
            header_row_ex = None
            for r_ in range(1, min(ws_ex.max_row + 1, 10)):
                v1 = str(ws_ex.cell(r_, 1).value or "").strip()
                v2 = str(ws_ex.cell(r_, 2).value or "").strip()
                v3 = str(ws_ex.cell(r_, 3).value or "").strip()
                if v1 == "Contor" and v2 == "An" and v3 == "Luna":
                    header_row_ex = r_
                    break

            if header_row_ex:
                ex_headers = {}
                for c in range(1, ws_ex.max_column + 1):
                    v = ws_ex.cell(header_row_ex, c).value
                    if v is not None:
                        ex_headers[str(v).strip()] = c

                rows_ex = []
                for r_ in range(header_row_ex + 1, ws_ex.max_row + 1):
                    ct = ws_ex.cell(r_, 1).value
                    an = ws_ex.cell(r_, 2).value
                    lu = ws_ex.cell(r_, 3).value
                    if ct is None or an is None:
                        continue
                    try:
                        an_int = int(float(str(an)))
                        lu_int = int(float(str(lu))) if lu is not None else 0
                    except Exception:
                        continue
                    if not (2000 <= an_int <= 2100):
                        continue
                    if not (1 <= lu_int <= 12):
                        continue
                    if str(ct).strip() in site_ids:
                        continue  # va fi rescris cu datele noi din batch

                    row_dict = {}
                    for col_name, col_idx in ex_headers.items():
                        row_dict[col_name] = ws_ex.cell(r_, col_idx).value
                    rows_ex.append(row_dict)

                wb_ex.close()
                df_final = pd.concat(
                    [pd.DataFrame(rows_ex), df_new], ignore_index=True
                ) if rows_ex else df_new
            else:
                wb_ex.close()
                df_final = df_new
        except Exception:
            df_final = df_new
    else:
        df_final = df_new

    # ── 3. Sortare ────────────────────────────────────────────────────────────
    for col in ['An', 'Luna']:
        if col in df_final.columns:
            df_final[col] = pd.to_numeric(df_final[col], errors='coerce')
    df_final = df_final.sort_values(
        [c for c in ['Contor', 'An', 'Luna'] if c in df_final.columns]
    ).reset_index(drop=True)

    # ── 4+. Scriere efectivă — O SINGURĂ DATĂ pentru tot batch-ul ──────────────
    _write_centralizator_file(df_final, _last_user)

    return CENTRAL_FILE, erori


def update_centralizator(excel_path, site_id, central_folder):
    """
    Wrapper de compatibilitate — actualizează centralizatorul pentru UN
    SINGUR contor (rescrie fișierul integral, ca înainte).

    Pentru procesare în masă (mai multe contoare într-o rulare), folosiți
    update_centralizator_batch() — face o singură rescriere pentru tot
    batch-ul în loc de una per contor, ceea ce reduce drastic timpul total
    (rescrierea completă, cu grafice și backup, e costul dominant, nu citirea
    datelor).
    """
    _central_file, erori = update_centralizator_batch(
        [{"path": excel_path, "id": site_id}], central_folder)
    if erori:
        raise erori[0][1]
    return _central_file

