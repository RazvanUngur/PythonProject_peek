# =============================================================================
# contoare_db.py — Gestionare bază de date contoare (JSON local)
# =============================================================================
# Exportă:
#   _load_contoare_db()
#   _save_contoare_db(db)
#   _delete_contor_from_centralizator(ct_id)
# =============================================================================

import os
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    CENTRAL_FILE_NAME, CENTRAL_FILE_FOLDER,
    CONTOARE_COLS_ORDER, CONTOARE_HEADERS, TIP_OPTIONS,
)

# ══════════════════════════════════════════════════════════════════════════════
# GESTIONARE BAZĂ DE DATE CONTOARE
# ══════════════════════════════════════════════════════════════════════════════



def _load_contoare_db():
    """
    Citește sheet-ul 'Contoare' din centralizator și returnează dict:
    { "2168": {"Drum": "DN1", "Pozitie_km": "km 45", "Localitate": "...", "Tip": "...", "IP": "..."} }
    """
    db = {}
    central_file = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)
    if not os.path.exists(central_file):
        return db
    try:
        wb = openpyxl.load_workbook(central_file, data_only=True)
        if "Contoare" not in wb.sheetnames:
            wb.close()
            return db
        ws = wb["Contoare"]
        # Citim header rândul 1
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v is not None:
                headers[str(v).strip()] = c

        for r in range(2, ws.max_row + 1):
            ct_val = ws.cell(r, 1).value
            if ct_val is None:
                continue
            ct_key = str(ct_val).strip()
            entry = {}
            # Mapeaza headerele vechi si noi
            col_map = {
                "Drum":          headers.get("Drum"),
                "Pozitie_km":    headers.get("Poziție km") or headers.get("Pozitie km") or headers.get("Poziție km."),
                "Localitate":    headers.get("Localitate"),
                "Tip":           headers.get("Tip"),
                "IP":            headers.get("IP"),
            }
            for field, col_idx in col_map.items():
                if col_idx is not None:
                    v = ws.cell(r, col_idx).value
                    entry[field] = str(v).strip() if v is not None else ""
                else:
                    entry[field] = ""
            db[ct_key] = entry
        wb.close()
    except Exception as e:
        print(f"[WARN] _load_contoare_db: {e}")
    return db


def _delete_contor_from_centralizator(ct_id):
    """
    Șterge toate datele contorului ct_id din centralizator:
      - Sheet 'Contoare'    : rândul cu Contor == ct_id
      - Sheet 'Media Zilnica Lunara': toate rândurile cu coloana Contor == ct_id
    """
    central_file = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)
    if not os.path.exists(central_file):
        return

    ct_id_str = str(ct_id)
    wb = openpyxl.load_workbook(central_file)

    # ── Sheet 'Contoare' ──────────────────────────────────────────────────────
    if "Contoare" in wb.sheetnames:
        ws_ct = wb["Contoare"]
        rows_to_delete = []
        for row in ws_ct.iter_rows(min_row=2):
            if str(row[0].value) == ct_id_str:
                rows_to_delete.append(row[0].row)
        # Ștergem de jos în sus ca să nu decalăm indicii
        for r in sorted(rows_to_delete, reverse=True):
            ws_ct.delete_rows(r)

    # ── Sheet 'Media Zilnica Lunara' ──────────────────────────────────────────
    if "Media Zilnica Lunara" in wb.sheetnames:
        ws_mz = wb["Media Zilnica Lunara"]
        # Găsim coloana 'Contor' (poate fi pe rândul 1 sau 2 dacă există titlu)
        header_row = None
        contor_col = None
        for r in range(1, 4):
            for c in range(1, ws_mz.max_column + 1):
                if str(ws_mz.cell(r, c).value).strip().lower() == "contor":
                    header_row = r
                    contor_col = c
                    break
            if contor_col:
                break

        if contor_col:
            rows_to_delete = []
            for row in ws_mz.iter_rows(min_row=header_row + 1):
                if str(row[contor_col - 1].value) == ct_id_str:
                    rows_to_delete.append(row[0].row)
            for r in sorted(rows_to_delete, reverse=True):
                ws_mz.delete_rows(r)

    wb.save(central_file)


def _save_contoare_db(db):
    """
    Scrie dict-ul db în sheet-ul 'Contoare' al centralizatorului.
    Dacă fișierul nu există, îl creează cu ambele sheet-uri necesare.
    """
    os.makedirs(CENTRAL_FILE_FOLDER, exist_ok=True)
    central_file = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)

    C_DARK = "1F4E79"; C_MID = "2E75B6"; C_LIGHT = "D6E4F0"
    C_WHITE = "FFFFFF"
    thin = Side(style='thin', color='BFBFBF')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    paleta = [C_LIGHT, "E8F4FD", "EBF5EB", "FDF6E3", "F5EEF8"]

    def hfont(sz=10):
        return Font(name='Arial', size=sz, bold=True, color=C_WHITE)

    def dfont(sz=10, bold=False):
        return Font(name='Arial', size=sz, bold=bold, color="1F1F1F")

    def fill(hex_c):
        return PatternFill('solid', start_color=hex_c)

    def ctr():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    if os.path.exists(central_file):
        wb = openpyxl.load_workbook(central_file)
        # Sterge sheet-ul vechi Contoare și îl recreează
        if "Contoare" in wb.sheetnames:
            del wb["Contoare"]
        ws = wb.create_sheet("Contoare", 1)  # inserează după primul sheet
    else:
        wb = openpyxl.Workbook()
        # Redenumim sheet-ul default
        wb.active.title = "Media Zilnica Lunara"
        ws = wb.create_sheet("Contoare", 1)

    ws.sheet_view.showGridLines = False

    # Header
    col_widths = [16, 22, 14, 22, 26, 18]
    for c_idx, hdr in enumerate(CONTOARE_HEADERS, 1):
        cell = ws.cell(1, c_idx, hdr)
        cell.font      = hfont(10)
        cell.fill      = fill(C_DARK if c_idx == 1 else C_MID)
        cell.alignment = ctr()
        cell.border    = brd
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths[c_idx - 1]
    ws.row_dimensions[1].height = 26

    # Date
    for r_idx, (ct_id, data) in enumerate(sorted(db.items()), 2):
        row_fill = fill(paleta[(r_idx - 2) % len(paleta)])
        values = [
            ct_id,
            data.get("Drum", ""),
            data.get("Pozitie_km", ""),
            data.get("Localitate", ""),
            data.get("Tip", ""),
            data.get("IP", ""),
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(r_idx, c_idx, val)
            cell.font      = dfont(10, bold=(c_idx == 1))
            cell.fill      = row_fill
            cell.alignment = ctr()
            cell.border    = brd

    ws.freeze_panes = "A2"
    wb.save(central_file)


