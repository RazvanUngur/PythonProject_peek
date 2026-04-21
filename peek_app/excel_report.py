# =============================================================================
# excel_report.py — Generare rapoarte Excel (foi, grafice, formatare)
# =============================================================================
# Exportă:
#   add_charts_and_formatting(excel_path, df, site_id)
# =============================================================================

import os
import calendar
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    VEHICLE_ANALYSIS, MIN_LUNI_AN, MIN_LUNI_AN_MAI, MIN_LUNI_AN_OCT,
    MIN_ORE_ZI, BAND_COLORS,
    CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME,
)

def add_charts_and_formatting(excel_path, df, site_id, source_files=None, source_file_periods=None):
    MIN_ORE_ZI    = 22
    MIN_ZILE_LUNA = 15
    MIN_ZILE_SAPT = 7
    # importăm globalele — deja definite la nivel de modul
    _MIN_LUNI_AN     = MIN_LUNI_AN
    _MIN_LUNI_AN_MAI = MIN_LUNI_AN_MAI
    _MIN_LUNI_AN_OCT = MIN_LUNI_AN_OCT

    C_DARK   = "1F4E79"
    C_MID    = "2E75B6"
    C_LIGHT  = "D6E4F0"
    C_WHITE  = "FFFFFF"
    C_GREEN  = "E2EFDA"
    C_ORANGE = "FCE4D6"
    C_YELLOW = "FFF2CC"
    C_VIOLET = "EAD1DC"   # violet pal — Prelucrare manuală MZL

    # Culori benzi
    BAND_COLORS = ["2E75B6", "ED7D31", "70AD47", "FFC000", "5B9BD5", "FF0000"]

    def fill(hex_c):
        return PatternFill('solid', start_color=hex_c)

    thin = Side(style='thin', color='BFBFBF')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hfont(sz=10):
        return Font(name='Arial', size=sz, bold=True, color=C_WHITE)

    def dfont(sz=10, bold=False, color="1F1F1F"):
        return Font(name='Arial', size=sz, bold=bold, color=color)

    def ctr():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def rgt():
        return Alignment(horizontal='right', vertical='center')

    days_ro = ['Luni', 'Marți', 'Miercuri', 'Joi', 'Vineri', 'Sâmbătă', 'Duminică']

    wb = load_workbook(excel_path)

    df = df.copy()  # defragmentare DataFrame (evită PerformanceWarning)
    df['Data'] = pd.to_datetime(df['Data_Ora'], format='%d.%m.%Y %H:%M')
    df['Zi']   = df['Data'].dt.date
    df['Ora']  = df['Data'].dt.hour
    df['An']   = df['Data'].dt.year
    df['Luna'] = df['Data'].dt.month

    # ── Detectare dinamică număr de benzi ───────────────────────────────────
    # Numărăm câte coloane Total_BX există în DataFrame
    n_bands = sum(1 for b in range(1, 7) if f'Total_B{b}' in df.columns)
    if n_bands not in (2, 4, 6):
        n_bands = 2  # fallback sigur
    band_ids = list(range(1, n_bands + 1))

    # Sens 1 = prima jumătate de benzi, Sens 2 = a doua jumătate
    sens1_bands = band_ids[:n_bands // 2]
    sens2_bands = band_ids[n_bands // 2:]

    # Coloana Total_General = suma tuturor benzilor
    total_band_cols = [f'Total_B{b}' for b in band_ids]
    if 'Total_General' not in df.columns:
        df['Total_General'] = df[total_band_cols].sum(axis=1)

    # Totaluri pe sensuri
    df['Total_Sens1'] = df[[f'Total_B{b}' for b in sens1_bands]].sum(axis=1)
    df['Total_Sens2'] = df[[f'Total_B{b}' for b in sens2_bands]].sum(axis=1)

    agg_dict = {col: 'sum' for col in total_band_cols}
    agg_dict['Total_General'] = 'sum'
    agg_dict['Total_Sens1'] = 'sum'
    agg_dict['Total_Sens2'] = 'sum'

    daily_data = df.groupby('Zi').agg(agg_dict).reset_index()
    for b in band_ids:
        daily_data[f'Peak_B{b}'] = df.groupby('Zi')[f'Total_B{b}'].max().values
    daily_data['Ore'] = df.groupby('Zi').size().values

    hourly_avg_dict = {f'Total_B{b}': 'mean' for b in band_ids}
    hourly_avg_dict['Total_Sens1'] = 'mean'
    hourly_avg_dict['Total_Sens2'] = 'mean'
    hourly_avg = df.groupby('Ora').agg(hourly_avg_dict).reset_index()

    n_days         = len(daily_data)
    latime_dinamica = max(20, n_days * 0.6)

    # ── FOAIE 2: Rezumat Zilnic ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Rezumat Zilnic")
    ws2.freeze_panes = "C3"

    ore_pe_zi = df.groupby('Zi').size().reset_index(name='Ore_pe_zi')

    # Coloane clase existente pentru fiecare bandă
    class_cols_all = {}
    for b in band_ids:
        class_cols_all[b] = [f'B{b}_Clasa_{i}' for i in range(1, 9)] + [f'B{b}_Clasa_15']

    agg_class = {}
    for b in band_ids:
        for col in class_cols_all[b]:
            if col in df.columns:
                agg_class[col] = 'sum'
    for b in band_ids:
        agg_class[f'Total_B{b}'] = 'sum'

    daily_classes = df.groupby('Zi').agg(agg_class).reset_index()
    daily_classes = daily_classes.merge(ore_pe_zi, on='Zi')

    # Suma clase per tip (toate benzile)
    for i in list(range(1, 9)) + [15]:
        cols_ci = [f'B{b}_Clasa_{i}' for b in band_ids if f'B{b}_Clasa_{i}' in daily_classes.columns]
        daily_classes[f'Clasa_{i}'] = daily_classes[cols_ci].sum(axis=1) if cols_ci else 0

    clasa_cols_for_total = [f'Clasa_{i}' for i in list(range(1, 9)) + [15]
                            if f'Clasa_{i}' in daily_classes.columns]
    daily_classes['Total'] = daily_classes[clasa_cols_for_total].sum(axis=1)

    mask_insuficient = daily_classes['Ore_pe_zi'] < 22
    zero_cols = [f'Clasa_{i}' for i in list(range(1, 9)) + [15]] + ['Total']
    for b in band_ids:
        zero_cols += [f'Total_B{b}']
    for col in zero_cols:
        if col in daily_classes.columns:
            daily_classes.loc[mask_insuficient, col] = 0

    for b in band_ids:
        daily_classes[f'Peak_B{b}'] = df.groupby('Zi')[f'Total_B{b}'].max().values
        daily_classes.loc[mask_insuficient, f'Peak_B{b}'] = 0

    first_date = daily_classes['Zi'].min().strftime('%d.%m.%Y')
    last_date  = daily_classes['Zi'].max().strftime('%d.%m.%Y')

    # --- Extragere Localitate din Centralizator ---
    localitate_site = ""
    try:
        path_centralizator = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)
        if os.path.exists(path_centralizator):
            df_contoare = pd.read_excel(path_centralizator, sheet_name='Contoare')
            match = df_contoare[df_contoare.iloc[:, 0].astype(str).str.contains(str(site_id))]
            if not match.empty:
                localitate_site = f" - {match.iloc[0, 3]}"
    except Exception as e:
        print(f"Atenție: Nu s-a putut citi localitatea din centralizator: {e}")

    # ── Construire headers dinamici pentru Rezumat Zilnic ───────────────────
    # Structura: Data | Zi | Clase 1-8 | Clasa 15 | TOTAL | Total_B1..BN | Varf_B1..BN | B1%..BN% | Indicator | Ore
    # 2 benzi: 2 + 9 + 1 + N + N + N + 2 = 2+9+1+2+2+2+2 = 20 col
    # 4 benzi: 2+9+1+4+4+4+2 = 26 col
    # 6 benzi: 2+9+1+6+6+6+2 = 32 col

    headers = ["Data", "Zi săptămână",
               "Clasa 1","Clasa 2","Clasa 3","Clasa 4","Clasa 5",
               "Clasa 6","Clasa 7","Clasa 8","Clasa 15",
               "TOTAL"]
    for b in band_ids:
        headers.append(f"Total Banda {b}")
    for b in band_ids:
        headers.append(f"Vârf Banda {b}")
    for b in band_ids:
        headers.append(f"Banda {b} %")
    headers += ["Indicator", "Ore înregistrate", "Verificare"]

    n_total_cols = len(headers)
    total_col_idx = 12  # coloana L (1-based) = TOTAL
    # Coloanele Total_BX: de la col 13 la 12+n_bands
    total_b_start_col = 13
    # Coloanele Varf_BX: de la col 13+n_bands la 12+2*n_bands
    varf_b_start_col = total_b_start_col + n_bands
    # Coloanele Proc_BX: de la col 13+2*n_bands la 12+3*n_bands
    proc_b_start_col = varf_b_start_col + n_bands
    # Indicator, Ore, Verificare
    indicator_col  = proc_b_start_col + n_bands
    ore_col        = indicator_col + 1
    verificare_col = ore_col + 1          # coloana nouă „Verificare"

    # Titlu dinamic care acoperă toate coloanele
    title_end_col = get_column_letter(n_total_cols)
    ws2.merge_cells(f"A1:{title_end_col}1")
    ws2["A1"] = f"REZUMAT ZILNIC  |  Contor: {site_id}{localitate_site}  |  {first_date} – {last_date}"
    ws2["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws2["A1"].fill = fill(C_DARK)
    ws2["A1"].alignment = ctr()
    ws2.row_dimensions[1].height = 28

    for c, h in enumerate(headers, 1):
        cell = ws2.cell(2, c, h)
        cell.font = hfont(9)
        # Coloana Verificare — header distinct (violet închis)
        if h == "Verificare":
            cell.fill = fill("5B2C6F")
        else:
            cell.fill = fill(C_MID)
        cell.alignment = ctr()
        cell.border = brd
    ws2.row_dimensions[2].height = 32

    # Calculăm ore cu trafic REAL
    ore_cu_trafic = (
        df[df['Total_General'] > 0]
        .groupby('Zi')
        .size()
        .reset_index(name='Ore_cu_trafic')
    )
    daily_classes = daily_classes.merge(ore_cu_trafic, on='Zi', how='left')
    daily_classes['Ore_cu_trafic'] = daily_classes['Ore_cu_trafic'].fillna(0).astype(int)

    dr = 3
    for i, row in daily_classes.iterrows():
        dt = row['Zi']
        day_name = days_ro[dt.weekday()]
        ore_cu_trafic_zi = int(row['Ore_cu_trafic'])

        if ore_cu_trafic_zi == 0:
            indicator = "Nu există date"
            rf = fill(C_ORANGE)
        elif ore_cu_trafic_zi < MIN_ORE_ZI:
            indicator = "Date parțiale"
            rf = fill(C_YELLOW)
            for col in zero_cols:
                if col in row.index:
                    row[col] = 0
        elif ore_cu_trafic_zi == 24:
            indicator = "Date complete"
            rf = fill(C_LIGHT) if i % 2 == 0 else fill(C_WHITE)
        else:
            indicator = "Date parțiale"
            rf = fill(C_YELLOW)

        # Construire valori rând
        vals = [dt.strftime("%d.%m.%Y"), day_name]
        for ci in list(range(1, 9)) + [15]:
            vals.append(int(row.get(f'Clasa_{ci}', 0)))
        vals.append(int(row.get('Total', 0)))  # TOTAL col 12
        total_col_letter = get_column_letter(total_col_idx)
        for b in band_ids:
            vals.append(int(row.get(f'Total_B{b}', 0)))
        for b in band_ids:
            vals.append(int(row.get(f'Peak_B{b}', 0)))
        for bi, b in enumerate(band_ids):
            b_col = get_column_letter(total_b_start_col + bi)
            vals.append(f"=IFERROR({b_col}{dr}/{total_col_letter}{dr},0)")
        vals.append(indicator)
        vals.append(f"{ore_cu_trafic_zi}/24")

        # ── Coloana „Verificare" — status per bandă (C / T / D) ─────────────
        # C = clasificator, T = totalizator (Clasa_15 > 10%), D = defect (0 veh)
        PRAG_TOT = 0.10
        verificare_parts = []
        for b in band_ids:
            total_b_zi = int(row.get(f'Total_B{b}', 0))
            cls15_b_zi = int(row.get(f'B{b}_Clasa_15', 0))
            if total_b_zi == 0:
                verificare_parts.append("D")
            elif cls15_b_zi / total_b_zi > PRAG_TOT:
                verificare_parts.append("T")
            else:
                verificare_parts.append("C")
        verificare_text = ", ".join(verificare_parts)
        vals.append(verificare_text)

        for c, val in enumerate(vals, 1):
            cell = ws2.cell(dr, c, val)
            cell.border = brd
            cell.alignment = ctr() if (c <= 2 or c >= indicator_col) else rgt()
            if 3 <= c <= total_col_idx + n_bands * 2:
                cell.number_format = '#,##0'
            if proc_b_start_col <= c <= proc_b_start_col + n_bands - 1:
                cell.number_format = '0.0%'
            if ore_cu_trafic_zi == 0:
                cell.font = dfont(9, bold=True, color="C00000")
                cell.fill = fill(C_ORANGE)
            else:
                cell.font = dfont(9)
                cell.fill = rf

            # Suprascriere stil celulă Verificare
            if c == verificare_col:
                if ore_cu_trafic_zi == 0:
                    cell.font = dfont(9, bold=True, color="C00000")
                else:
                    has_T = "T" in verificare_parts
                    has_D = "D" in verificare_parts
                    has_C = "C" in verificare_parts
                    if has_D:
                        v_fill = "FADBD8"; v_color = "922B21"   # roz — defect
                    elif has_T and not has_C:
                        v_fill = "FCE4D6"; v_color = "C55A11"   # portocaliu — toate TOT
                    elif has_T:
                        v_fill = "FFF2CC"; v_color = "7D6608"   # galben — mix T+C
                    else:
                        v_fill = "D5F5E3"; v_color = "1E8449"   # verde — toate CLS
                    cell.font = dfont(9, bold=True, color=v_color)
                    cell.fill = fill(v_fill)
        dr += 1

    last_row = dr - 1
    ws2.merge_cells(f"A{dr}:B{dr}")
    ws2[f"A{dr}"] = "TOTAL PERIOADĂ"
    ws2[f"A{dr}"].font = dfont(bold=True)
    ws2[f"A{dr}"].fill = fill(C_GREEN)
    ws2[f"A{dr}"].alignment = ctr()
    ws2[f"A{dr}"].border = brd
    ws2[f"B{dr}"].fill = fill(C_GREEN)
    ws2[f"B{dr}"].border = brd

    for c in range(3, n_total_cols + 1):
        cl = get_column_letter(c)
        is_varf = varf_b_start_col <= c <= varf_b_start_col + n_bands - 1
        is_proc = proc_b_start_col <= c <= proc_b_start_col + n_bands - 1
        is_text = c >= indicator_col
        if is_varf:
            formula = f"=MAX({cl}3:{cl}{last_row})"
        elif is_proc:
            b_offset = c - proc_b_start_col
            b_col = get_column_letter(total_b_start_col + b_offset)
            formula = f"=IFERROR({b_col}{dr}/{get_column_letter(total_col_idx)}{dr},0)"
        elif is_text:
            formula = ""
        else:
            formula = f"=SUM({cl}3:{cl}{last_row})"
        cell = ws2.cell(dr, c, formula)
        cell.font = dfont(bold=True)
        cell.fill = fill(C_GREEN)
        cell.alignment = rgt() if not is_text else ctr()
        cell.border = brd
        if 3 <= c <= varf_b_start_col + n_bands - 1:
            cell.number_format = '#,##0'
        elif is_proc:
            cell.number_format = '0.0%'

    widths = [12, 13] + [9] * 9 + [11] + [13] * n_bands + [13] * n_bands + [11] * n_bands + [14, 14, 22]
    for c, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    def make_bar_layout():
        return Layout(manualLayout=ManualLayout(x=0.05, y=0.10, h=0.75, w=0.90,
                                                xMode="edge", yMode="edge"))

    def configure_x_axis(ax):
        ax.delete = False
        ax.axPos  = "b"
        ax.tickLblPos  = "nextTo"
        ax.textRotation = -45000
        ax.tickLblSkip  = 1

    # ── Pregătire date pentru Pie Charts — ultimele 30 zile valide ──────────
    dc_valide = daily_classes[~mask_insuficient].copy()
    dc_pie    = dc_valide.tail(30)

    clase_pie = [1, 2, 3, 4, 5, 6, 7, 8, 15]
    label_pie = [f"Clasa {i}" for i in clase_pie]

    # Sume pe clase pentru fiecare bandă
    sume_per_banda = {}
    for b in band_ids:
        sume_b = []
        for i in clase_pie:
            col = f'B{b}_Clasa_{i}'
            sume_b.append(int(dc_pie[col].sum()) if col in dc_pie.columns else 0)
        sume_per_banda[b] = sume_b

    # Plasăm datele auxiliare pentru pie charts
    # Coloana start: după ultimele coloane principale (n_total_cols + 2)
    PIE_AUX_START_COL = n_total_cols + 2  # lăsăm un spațiu
    PIE_START_ROW = 3
    n_clase = len(clase_pie)

    # Header etichete (comune)
    lbl_col = PIE_AUX_START_COL
    ws2.cell(PIE_START_ROW - 1, lbl_col, "Clasa").font = dfont(bold=True)
    for b in band_ids:
        val_col = PIE_AUX_START_COL + b
        ws2.cell(PIE_START_ROW - 1, val_col, f"Banda {b}").font = dfont(bold=True)

    for idx, lbl in enumerate(label_pie):
        r = PIE_START_ROW + idx
        ws2.cell(r, lbl_col, lbl)
        for b in band_ids:
            ws2.cell(r, PIE_AUX_START_COL + b, sume_per_banda[b][idx])

    # Culori fixe pentru clase
    CULORI_PIE = {
        0: "A9CCE3", 1: "27AE60", 2: "F39C12", 3: "8E44AD",
        4: "87CEFA", 5: "1ABC9C", 6: "2E75B6", 7: "F7DC6F", 8: "C0392B",
    }

    def make_pie(title, val_col, n):
        pie = PieChart()
        pie.title  = title
        pie.style  = 10
        pie.width  = 16
        pie.height = 16

        data = Reference(ws2, min_col=val_col, min_row=PIE_START_ROW - 1,
                         max_row=PIE_START_ROW + n - 1)
        cats = Reference(ws2, min_col=lbl_col,  min_row=PIE_START_ROW,
                         max_row=PIE_START_ROW + n - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(cats)

        serie = pie.series[0]
        for idx_c in range(n):
            pt = DataPoint(idx=idx_c)
            hex_c = CULORI_PIE.get(idx_c, "AAAAAA")
            pt.graphicalProperties.solidFill = hex_c
            pt.graphicalProperties.line.solidFill = "FFFFFF"
            serie.dPt.append(pt)

        serie.dLbls = DataLabelList()
        serie.dLbls.showPercent  = True
        serie.dLbls.showCatName  = False
        serie.dLbls.showVal      = False
        serie.dLbls.showSerName  = False
        serie.dLbls.showLeaderLines = True

        pie.legend.position = 'b'
        pie.legend.overlay  = False

        pie.layout = Layout(manualLayout=ManualLayout(
            x=0.05, y=0.20, h=0.65, w=0.90, xMode="edge", yMode="edge"
        ))
        return pie

    # Poziționare pie charts — pozitii fixe exacte per număr de benzi
    # 2 benzi: Banda1=V3,  Banda2=AH3  (ambele pe rândul 1)
    # 4 benzi: Banda1=AB3, Banda2=AL3, Banda3=AB35, Banda4=AL35
    # 6 benzi: Banda1=AH3, Banda2=AR3, Banda3=BB3,  Banda4=AH35, Banda5=AR35, Banda6=BB35

    if n_bands == 2:
        pie_anchors = ["V3", "AF3"]
        bar_anchor  = "V36"
    elif n_bands == 4:
        pie_anchors = ["AB3", "AL3", "AB35", "AL35"]
        bar_anchor  = "AB67"
    else:  # 6 benzi
        pie_anchors = ["AH3", "AR3", "BB3", "AH35", "AR35", "BB35"]
        bar_anchor  = "AH67"

    for b, anchor in zip(band_ids, pie_anchors):
        val_col   = PIE_AUX_START_COL + b
        pie_chart = make_pie(f"Distribuție clase — Banda {b}  (ult. 30 zile)", val_col, n_clase)
        ws2.add_chart(pie_chart, anchor)

    chart2 = BarChart()
    chart2.type = "col"; chart2.grouping = "percentStacked"; chart2.overlap = 100
    chart2.title = f"Distribuție procentuală {'  /  '.join([f'Banda {b}' for b in band_ids])}"
    chart2.y_axis.title = "Procent (%)"; chart2.style = 10
    chart2.width = latime_dinamica; chart2.height = 12
    chart2.legend.position = 'b'; chart2.legend.overlay = False
    configure_x_axis(chart2.x_axis)
    chart2.layout = make_bar_layout()

    # Date pentru chart: coloanele Total_BX din ws2
    for bi, b in enumerate(band_ids):
        col_idx = total_b_start_col + bi
        ref = Reference(ws2, min_col=col_idx, min_row=2, max_row=2 + n_days)
        chart2.add_data(ref, titles_from_data=True)
        chart2.series[bi].graphicalProperties.solidFill = BAND_COLORS[bi % len(BAND_COLORS)]

    cats_all = Reference(ws2, min_col=1, min_row=3, max_row=2 + n_days)
    chart2.set_categories(cats_all)
    ws2.add_chart(chart2, bar_anchor)

    # ── FOAIE 3: Date prelucrate (după Rezumat Zilnic) ────────────────────────
    # Perechile bandă ↔ bandă sens opus:
    #   2 benzi: B1↔B2 | 4 benzi: B1↔B4, B2↔B3 | 6 benzi: B1↔B6, B2↔B5, B3↔B4
    ws_dp = wb.create_sheet("Date prelucrate")
    ws_dp.sheet_view.showGridLines = False

    CLASE_IDX = list(range(1, 9)) + [15]
    PERECHI = {2: {1:2,2:1}, 4: {1:4,2:3,3:2,4:1}, 6: {1:6,2:5,3:4,4:3,5:2,6:1}}
    perechi = PERECHI.get(n_bands, PERECHI[2])
    PRAG_TOT_DP = 0.10

    def _sb(tot_b, cls15_b):
        if tot_b == 0: return 'D'
        if cls15_b / tot_b > PRAG_TOT_DP: return 'T'
        return 'C'

    def _recon(donor_cls, donor_tot, target_tot):
        """
        Distribuie target_tot pe clase folosind proporțiile din donor.
        Dacă donor_cls[c] == 0 → rezultat[c] == 0 întotdeauna.
        Fiecare valoare e ≥ 0; ultimul element absoarbe restul de rotunjire,
        dar nu poate deveni negativ.
        """
        if donor_tot == 0 or target_tot == 0:
            return {c: 0 for c in CLASE_IDX}
        res = {}
        assigned = 0
        # Sortăm descrescător ca să distribuim primele clasele mari
        # (minimizează eroarea de rotunjire pe ultimul element)
        sorted_cls = sorted(CLASE_IDX,
                            key=lambda c: donor_cls.get(c, 0), reverse=True)
        for i, c in enumerate(sorted_cls):
            donor_val = donor_cls.get(c, 0)
            if donor_val == 0:
                res[c] = 0   # dacă donorul e 0 → rezultatul e 0, fără excepție
            elif i == len(sorted_cls) - 1:
                # Ultimul element: restul, dar minim 0
                res[c] = max(0, target_tot - assigned)
            else:
                v = round(donor_val / donor_tot * target_tot)
                res[c] = max(0, v)   # niciodată negativ
                assigned += res[c]
        return res

    def _copy(src): return {c: max(0, src.get(c, 0)) for c in CLASE_IDX}

    def _scale(src_cls, src_tot, factor):
        """Scalează clasele src cu factorul factor/src_tot. Valori ≥ 0."""
        if src_tot == 0 or factor <= 0:
            return {c: 0 for c in CLASE_IDX}
        res = {}
        assigned = 0
        sorted_cls = sorted(CLASE_IDX,
                            key=lambda c: src_cls.get(c, 0), reverse=True)
        for i, c in enumerate(sorted_cls):
            src_val = src_cls.get(c, 0)
            if src_val == 0:
                res[c] = 0
            elif i == len(sorted_cls) - 1:
                res[c] = max(0, factor - assigned)
            else:
                v = round(src_val / src_tot * factor)
                res[c] = max(0, v)
                assigned += res[c]
        return res

    # Agregare zilnică
    _df_dp = df.copy()
    _df_dp['_zi'] = pd.to_datetime(_df_dp['Data_Ora'], format='%d.%m.%Y %H:%M',
                                    errors='coerce').dt.date
    _agg_dp = {}
    for _ci in CLASE_IDX:
        for _b in band_ids:
            _col = f'B{_b}_Clasa_{_ci}'
            if _col in _df_dp.columns: _agg_dp[_col] = 'sum'
    for _b in band_ids: _agg_dp[f'Total_B{_b}'] = 'sum'
    _daily_dp = _df_dp.groupby('_zi').agg(_agg_dp).reset_index().sort_values('_zi').reset_index(drop=True)

    SENS1_B = band_ids[:n_bands // 2]
    SENS2_B = band_ids[n_bands // 2:]

    # Coloane: Data|Zi|S1_cls..|S1_Total|S2_cls..|S2_Total|Total|Status benzi|Tip zi
    _dp_hdrs = ["Data", "Zi"]
    for _s, _sb_list in enumerate([SENS1_B, SENS2_B], 1):
        for _cls in CLASE_IDX: _dp_hdrs.append(f"S{_s}_Cls{_cls}")
        _dp_hdrs.append(f"S{_s}_Total")
    _dp_hdrs += ["Total", "Status benzi", "Tip zi"]
    _n_dp = len(_dp_hdrs)
    _title_col_dp = get_column_letter(_n_dp)

    # Rând 1 titlu — întreaga lățime
    ws_dp.merge_cells(f"A1:{_title_col_dp}1")
    ws_dp["A1"] = (f"DATE PRELUCRATE  |  Contor: {site_id}"
                   + (f"  |  {localitate_site.lstrip(' -')}" if localitate_site else "")
                   + "  |  Reconstrucție benzi zilnică")
    ws_dp["A1"].font      = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws_dp["A1"].fill      = fill(C_DARK)
    ws_dp["A1"].alignment = ctr()
    ws_dp.row_dimensions[1].height = 28

    # Rând 2 legendă — întreaga lățime
    ws_dp.merge_cells(f"A2:{_title_col_dp}2")
    ws_dp["A2"] = ("🟢 Verde=clasificator(real)   🔵 Albastru=reconstituit din pereche   "
                   "🟡 Galben=totalizator   🔴 Roșu=neutilizabil")
    ws_dp["A2"].font      = Font(name='Arial', size=8, italic=True, color="444444")
    ws_dp["A2"].fill      = fill("F8F8F8")
    ws_dp["A2"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_dp.row_dimensions[2].height = 20

    # Rând 3 header
    for _ci, _h in enumerate(_dp_hdrs, 1):
        _cell = ws_dp.cell(3, _ci, _h)
        _cell.border = brd; _cell.alignment = ctr()
        if _h in ("Data","Zi","Status benzi","Tip zi"):
            _cell.font = hfont(9); _cell.fill = fill(C_DARK)
        elif _h == "Total":
            _cell.font = hfont(9); _cell.fill = fill("404040")
        elif _h.startswith("S1_"):
            _cell.font = hfont(9); _cell.fill = fill("1F4E79")
        else:
            _cell.font = hfont(9); _cell.fill = fill("375623")
    ws_dp.row_dimensions[3].height = 26

    # Lățimi
    ws_dp.column_dimensions['A'].width = 12
    ws_dp.column_dimensions['B'].width = 12
    for _ci in range(3, _n_dp - 2):
        ws_dp.column_dimensions[get_column_letter(_ci)].width = 9
    ws_dp.column_dimensions[get_column_letter(_n_dp - 2)].width = 13  # Total
    ws_dp.column_dimensions[get_column_letter(_n_dp - 1)].width = 20  # Status
    ws_dp.column_dimensions[get_column_letter(_n_dp)].width = 14      # Tip zi

    F_REAL = "D9EAD3"; F_RECON = "CFE2F3"; F_TOT = "FFF2CC"
    F_NULL = "F4CCCC"; F_ALT   = "F8F9FA"
    _days_ro = ['Luni','Marți','Miercuri','Joi','Vineri','Sâmbătă','Duminică']

    # Dict tip_zi per zi — folosit de centralizator pentru Mod funcționare
    _dp_tip_zi_map = {}

    _dp_row = 4
    for _, _rz in _daily_dp.iterrows():
        _zi_date = _rz['_zi']
        _stat = {}; _rcls = {}; _rtot = {}
        for _b in band_ids:
            _tot = int(_rz.get(f'Total_B{_b}', 0))
            _c15 = int(_rz.get(f'B{_b}_Clasa_15', 0))
            _stat[_b] = _sb(_tot, _c15)
            _rtot[_b] = _tot
            _rcls[_b] = {_c: int(_rz.get(f'B{_b}_Clasa_{_c}', 0)) for _c in CLASE_IDX}

        _status_str = ', '.join(_stat[_b] for _b in band_ids)
        _ocls = {_b: {} for _b in band_ids}
        _otot = {_b: 0  for _b in band_ids}
        _osrc = {_b: 'D' for _b in band_ids}
        _tip_zi = 'clasificator'; _null_zi = False

        for _b in band_ids:
            _pb  = perechi[_b]
            _sb_ = _stat[_b]; _spb = _stat[_pb]

            if _sb_ == 'C':
                _ocls[_b] = _copy(_rcls[_b]); _otot[_b] = _rtot[_b]; _osrc[_b] = 'C'

            elif _sb_ == 'T':
                if _spb == 'C' and _rtot[_pb] > 0:
                    _ocls[_b] = _recon(_rcls[_pb], _rtot[_pb], _rtot[_b])
                    _otot[_b] = sum(_ocls[_b].values()); _osrc[_b] = 'R'
                else:
                    _ocls[_b] = {_c: 0 for _c in CLASE_IDX}
                    _otot[_b] = _rtot[_b]; _osrc[_b] = 'T'; _tip_zi = 'totalizator'

            else:  # D
                if _spb == 'C':
                    # Bandă defectă, pereche clasificatoare → preia valorile perechii
                    _ss_pb = SENS1_B if _pb in SENS1_B else SENS2_B
                    _ss_b  = SENS1_B if _b in SENS1_B else SENS2_B
                    _fip   = [_x for _x in _ss_pb if _x != _pb and _stat[_x] == 'C']
                    if _fip:
                        _bfs = [_x for _x in _ss_b if _stat[_x] == 'C']
                        if _bfs:
                            _db = _bfs[0]; _rb = perechi[_db]
                            if _rtot[_rb] > 0:
                                _sc = (_rtot[_pb] / _rtot[_rb]) * _rtot[_db]
                                _ocls[_b] = _scale(_rcls[_db], _rtot[_db], round(_sc))
                                _otot[_b] = sum(_ocls[_b].values()); _osrc[_b] = 'R'
                            else:
                                _ocls[_b] = _copy(_rcls[_pb]); _otot[_b] = _rtot[_pb]; _osrc[_b] = 'R'
                        else:
                            _ocls[_b] = _copy(_rcls[_pb]); _otot[_b] = _rtot[_pb]; _osrc[_b] = 'R'
                    else:
                        _ocls[_b] = _copy(_rcls[_pb]); _otot[_b] = _rtot[_pb]; _osrc[_b] = 'R'

                elif _spb == 'T':
                    # Pereche totalizatoare → preia totalul perechii
                    _ocls[_b] = {_c: 0 for _c in CLASE_IDX}
                    _otot[_b] = _rtot[_pb]; _osrc[_b] = 'T'; _tip_zi = 'totalizator'

                else:  # pb == D, ambele defecte → caută alt sens
                    _ss_b = SENS1_B if _b in SENS1_B else SENS2_B
                    _af = [_x for _x in _ss_b if _x != _b and _stat[_x] in ('C','T')]
                    if _af:
                        _xb = _af[0]; _xpb = perechi[_xb]
                        if _stat[_xb] == 'C' and _stat[_xpb] == 'C' and _rtot[_xb] > 0:
                            _ocls[_b] = _copy(_rcls[_xb]); _otot[_b] = _rtot[_xb]; _osrc[_b] = 'R'
                        else:
                            _null_zi = True; _ocls[_b] = {_c:0 for _c in CLASE_IDX}; _otot[_b] = 0; _osrc[_b] = 'N'
                    else:
                        _null_zi = True; _ocls[_b] = {_c:0 for _c in CLASE_IDX}; _otot[_b] = 0; _osrc[_b] = 'N'

        if _null_zi: _tip_zi = 'null'
        _dp_tip_zi_map[_zi_date] = _tip_zi

        # Scriere rând
        import datetime as _dtm
        _zi_obj  = _zi_date if hasattr(_zi_date,'weekday') else _dtm.date.fromisoformat(str(_zi_date))
        _row_bg  = F_NULL if _tip_zi=='null' else F_TOT if _tip_zi=='totalizator' else (F_ALT if _dp_row%2==0 else "FFFFFF")

        for _ci, _val in enumerate([_zi_obj.strftime('%d.%m.%Y'), _days_ro[_zi_obj.weekday()]], 1):
            _c = ws_dp.cell(_dp_row, _ci, _val)
            _c.font=dfont(9); _c.fill=fill(_row_bg); _c.border=brd; _c.alignment=ctr()

        _col = 3; _vs1 = 0; _vs2 = 0
        for _si, _sbn in enumerate([SENS1_B, SENS2_B]):
            for _cls in CLASE_IDX:
                _vc = sum(_ocls[_b].get(_cls,0) for _b in _sbn)
                _srcs = [_osrc[_b] for _b in _sbn]
                _bg = (F_NULL if _tip_zi=='null' or 'N' in _srcs else
                       F_TOT  if _tip_zi=='totalizator' else
                       F_RECON if 'R' in _srcs else F_REAL)
                _c = ws_dp.cell(_dp_row, _col, _vc if _tip_zi!='null' else "—")
                _c.font=dfont(9); _c.fill=fill(_bg); _c.border=brd; _c.alignment=rgt()
                if isinstance(_vc,int) and _tip_zi!='null': _c.number_format='#,##0'
                _col += 1
            _vst = sum(_otot[_b] for _b in _sbn)
            if _si==0: _vs1=_vst
            else: _vs2=_vst
            _bg_t = (F_NULL if _tip_zi=='null' else F_TOT if _tip_zi=='totalizator' else
                     F_RECON if any(_osrc[_b]=='R' for _b in _sbn) else F_REAL)
            _c = ws_dp.cell(_dp_row, _col, _vst if _tip_zi!='null' else "—")
            _c.font=dfont(9,bold=True); _c.fill=fill(_bg_t); _c.border=brd; _c.alignment=rgt()
            if isinstance(_vst,int) and _tip_zi!='null': _c.number_format='#,##0'
            _col += 1

        # Coloana Total (S1+S2)
        _vtg = _vs1 + _vs2
        _c = ws_dp.cell(_dp_row, _col, _vtg if _tip_zi!='null' else "—")
        _c.font=dfont(9,bold=True); _c.border=brd; _c.alignment=rgt()
        _c.fill=fill(F_NULL if _tip_zi=='null' else F_TOT if _tip_zi=='totalizator' else "E2EFDA")
        if isinstance(_vtg,int) and _tip_zi!='null': _c.number_format='#,##0'
        _col += 1

        # Status benzi
        _c = ws_dp.cell(_dp_row, _col, _status_str)
        _c.font=dfont(9); _c.fill=fill(_row_bg); _c.border=brd; _c.alignment=ctr()
        _col += 1

        # Tip zi
        _td = {'clasificator':'Clasificator','totalizator':'Totalizator','null':'Neutilizabil'}.get(_tip_zi,_tip_zi)
        _tc = {'clasificator':"1E8449",'totalizator':"7D6608",'null':"922B21"}.get(_tip_zi,"000000")
        _c = ws_dp.cell(_dp_row, _col, _td)
        _c.font=Font(name='Arial',size=9,bold=True,color=_tc)
        _c.fill=fill(_row_bg); _c.border=brd; _c.alignment=ctr()

        ws_dp.row_dimensions[_dp_row].height = 18
        _dp_row += 1

    # Rând TOTAL
    ws_dp.merge_cells(f"A{_dp_row}:B{_dp_row}")
    _tc2 = ws_dp[f"A{_dp_row}"]
    _tc2.value="TOTAL PERIOADĂ"; _tc2.font=dfont(bold=True)
    _tc2.fill=fill(C_GREEN); _tc2.alignment=ctr(); _tc2.border=brd
    ws_dp[f"B{_dp_row}"].fill=fill(C_GREEN); ws_dp[f"B{_dp_row}"].border=brd
    for _ci in range(3, _n_dp-1):
        _cl=get_column_letter(_ci)
        _c2=ws_dp.cell(_dp_row,_ci,f'=SUMIF({_cl}4:{_cl}{_dp_row-1},"<>—")')
        _c2.font=dfont(bold=True); _c2.fill=fill(C_GREEN)
        _c2.border=brd; _c2.alignment=rgt(); _c2.number_format='#,##0'
    for _ci in [_n_dp-1, _n_dp]:
        _c3=ws_dp.cell(_dp_row,_ci,""); _c3.fill=fill(C_GREEN); _c3.border=brd
    ws_dp.freeze_panes = "C4"
    ws_dp.row_dimensions[_dp_row].height = 22

    # ── FOAIE 4: Media Zilnică Lunară ────────────────────────────────────────
    ws_lunar = wb.create_sheet("Media Zilnica Lunara")

    # ── Citim suprascrierile manuale direct din SQLite (trafic.db) ────────────
    _manual_overrides = {}   # {(an, luna): valoare_mzl_totala}
    try:
        from database import get_traffic_db as _get_tdb
        _manual_overrides = _get_tdb().get_mzl_manual(str(site_id))
    except Exception as _e:
        # Fallback: citim din sheet-ul Excel al centralizatorului (compatibilitate)
        try:
            _central_file = os.path.join(CENTRAL_FILE_FOLDER, CENTRAL_FILE_NAME)
            if os.path.exists(_central_file):
                _wb_c = openpyxl.load_workbook(_central_file, data_only=True)
                if "Prelucrare manuala" in _wb_c.sheetnames:
                    _ws_pm = _wb_c["Prelucrare manuala"]
                    _pm_headers = {}
                    for _ci in range(1, _ws_pm.max_column + 1):
                        _v = _ws_pm.cell(2, _ci).value
                        if _v is not None:
                            _pm_headers[str(_v).strip()] = _ci
                    _site_str = str(site_id)
                    for _r in range(3, _ws_pm.max_row + 1):
                        _ct_col  = _pm_headers.get("Contor")
                        _an_col  = _pm_headers.get("An")
                        _lu_col  = _pm_headers.get("Luna")
                        _mzl_col = _pm_headers.get("MZL_Manual")
                        if not all([_ct_col, _an_col, _lu_col, _mzl_col]):
                            break
                        _ct_val = _ws_pm.cell(_r, _ct_col).value
                        if _ct_val is None or str(_ct_val).strip() != _site_str:
                            continue
                        try:
                            _an_v  = int(float(str(_ws_pm.cell(_r, _an_col).value)))
                            _lu_v  = int(float(str(_ws_pm.cell(_r, _lu_col).value)))
                            _mzl_v = float(str(_ws_pm.cell(_r, _mzl_col).value))
                            _manual_overrides[(_an_v, _lu_v)] = _mzl_v
                        except Exception:
                            continue
                _wb_c.close()
        except Exception as _e2:
            print(f"[WARN] Nu s-au putut citi suprascrierile manuale: {_e2}")

    ore_pe_zi2  = df.groupby(['An', 'Luna', 'Zi']).size().reset_index(name='Ore')
    zile_valide = ore_pe_zi2[ore_pe_zi2['Ore'] >= MIN_ORE_ZI][['An', 'Luna', 'Zi']]
    df_valid    = df.merge(zile_valide, on=['An', 'Luna', 'Zi'], how='inner')
    toate_lunile = df.groupby(['An', 'Luna']).size().reset_index()[['An', 'Luna']]

    def check_7_consecutive_days(df_all, an, luna):
        import datetime as dt_mod
        zile_in_luna = calendar.monthrange(an, luna)[1]
        df_luna_ore  = df_all[(df_all['An'] == an) & (df_all['Luna'] == luna)]
        ore_dict     = df_luna_ore.groupby(df_luna_ore['Data'].dt.date).size().to_dict()
        for start_day in range(1, zile_in_luna - 5):
            week_dates = []
            valid_week = True
            for offset in range(7):
                current_date = dt_mod.date(an, luna, start_day + offset)
                if ore_dict.get(current_date, 0) < MIN_ORE_ZI:
                    valid_week = False
                    break
                week_dates.append(current_date)
            if valid_week:
                return True, week_dates
        return False, []

    # ── Construim tabelul Date prelucrate agregat pe zi (folosit pentru MZL) ──
    # MZL se calculează din coloana "Total" a sheet-ului Date prelucrate,
    # adică din valorile reconstruite (benzi D/T corectate) per zi validă.
    # Zi validă = minim MIN_ORE_ZI ore înregistrate în ziua respectivă.

    # Refolosim _daily_dp și _dp_tip_zi_map construite în blocul Date prelucrate
    # (variabile locale în scope-ul funcției, calculate anterior)

    def _get_dp_total_zi(zi_date):
        """Returnează totalul din Date prelucrate pentru o zi dată."""
        try:
            _rz = _daily_dp[_daily_dp['_zi'] == zi_date]
            if _rz.empty:
                return 0
            # Recalculăm out_tot exact ca în blocul Date prelucrate
            _stat_b = {}
            _rtot_b = {}
            for _b in band_ids:
                _t = int(_rz.iloc[0].get(f'Total_B{_b}', 0))
                _c = int(_rz.iloc[0].get(f'B{_b}_Clasa_15', 0))
                _stat_b[_b] = 'D' if _t == 0 else ('T' if _c/_t > PRAG_TOT_DP else 'C')
                _rtot_b[_b] = _t
            # Total reconstituit = suma out_tot per bandă
            # Simplificare: folosim _dp_tip_zi_map și recalculăm totalul
            _tip = _dp_tip_zi_map.get(zi_date, 'null')
            if _tip == 'null':
                return 0
            # Suma tuturor benzilor după reconstrucție
            # Benzile D preiau valoarea perechii → total = total real
            total_reconst = 0
            for _b in band_ids:
                _pb = perechi.get(_b, _b)
                if _stat_b[_b] == 'C':
                    total_reconst += _rtot_b[_b]
                elif _stat_b[_b] == 'T':
                    total_reconst += _rtot_b[_b]
                else:  # D
                    total_reconst += _rtot_b.get(_pb, 0)
            return total_reconst
        except Exception:
            return 0

    lunar_data = []
    for _, row in toate_lunile.iterrows():
        an, luna = row['An'], row['Luna']
        df_luna_valida = df_valid[(df_valid['An'] == an) & (df_valid['Luna'] == luna)]
        nr_zile_valide = df_luna_valida['Zi'].nunique() if len(df_luna_valida) > 0 else 0
        zile_in_luna   = calendar.monthrange(an, luna)[1]

        if nr_zile_valide == zile_in_luna:
            indicator, color = "Date complete", "NORMAL"
            df_calcul = df_luna_valida; divisor = zile_in_luna
        elif nr_zile_valide >= MIN_ZILE_LUNA:
            indicator, color = "Date parțiale", "YELLOW"
            df_calcul = df_luna_valida; divisor = nr_zile_valide
        else:
            has_7, week_dates = check_7_consecutive_days(df, an, luna)
            if has_7:
                indicator, color = "Date parțiale - 7 zile", "YELLOW"
                df_calcul = df[df['Zi'].isin(week_dates)]; divisor = 7
            else:
                indicator, color = "Nu există date", "ORANGE"
                df_calcul = pd.DataFrame(); divisor = 0

        if divisor > 0 and not df_calcul.empty:
            # ── MZL din Date prelucrate — valorile reconstruite per clasă ──────
            # Colectăm zilele valide din luna curentă care există în _daily_dp
            zile_valide_luna = df_calcul['Zi'].unique() if 'Zi' in df_calcul.columns else []

            sume_dp = {f'Clasa_{i}': 0 for i in range(1, 9)}
            sume_dp['Clasa_15'] = 0
            n_zile_dp = 0

            for _zi_d in zile_valide_luna:
                # Găsim rândul din _daily_dp
                _dp_row_zi = _daily_dp[_daily_dp['_zi'] == _zi_d]
                if _dp_row_zi.empty:
                    continue
                _tip = _dp_tip_zi_map.get(_zi_d, 'null')
                if _tip == 'null':
                    continue   # zi neutilizabilă → excludem

                # Suma claselor reconstruite (din out_cls per toate benzile)
                # Reconstrucție rapidă pentru clase
                _rz   = _dp_row_zi.iloc[0]
                _stat_b = {}; _rcls_b = {}; _rtot_b = {}
                for _b in band_ids:
                    _t = int(_rz.get(f'Total_B{_b}', 0))
                    _c = int(_rz.get(f'B{_b}_Clasa_15', 0))
                    _stat_b[_b] = 'D' if _t == 0 else ('T' if _t > 0 and _c/_t > PRAG_TOT_DP else 'C')
                    _rtot_b[_b] = _t
                    _rcls_b[_b] = {_ci: int(_rz.get(f'B{_b}_Clasa_{_ci}', 0))
                                   for _ci in CLASE_IDX}

                for _cls_i in list(range(1, 9)) + [15]:
                    _col_key = f'Clasa_{_cls_i}'
                    _val_cls = 0
                    for _b in band_ids:
                        _pb = perechi.get(_b, _b)
                        if _stat_b[_b] == 'C':
                            _val_cls += max(0, _rcls_b[_b].get(_cls_i, 0))
                        elif _stat_b[_b] == 'T':
                            # T: distribuție proporțională din pereche dacă e C
                            if _stat_b.get(_pb, 'D') == 'C' and _rtot_b.get(_pb, 0) > 0:
                                _donor_cls_val = _rcls_b[_pb].get(_cls_i, 0)
                                if _donor_cls_val == 0:
                                    pass  # donorul 0 → rezultatul 0
                                else:
                                    _ratio = _donor_cls_val / _rtot_b[_pb]
                                    _val_cls += max(0, round(_ratio * _rtot_b[_b]))
                            # dacă pereche nu e C → 0 clase (totalizator pur)
                        else:  # D — preia clasele perechii direct
                            if _stat_b.get(_pb, 'D') == 'C':
                                _val_cls += max(0, _rcls_b[_pb].get(_cls_i, 0))
                    sume_dp[_col_key] += max(0, _val_cls)
                n_zile_dp += 1

            if n_zile_dp > 0:
                medii = {f'Clasa_{i}': round(sume_dp[f'Clasa_{i}'] / n_zile_dp)
                         for i in range(1, 9)}
                medii['Clasa_15'] = round(sume_dp['Clasa_15'] / n_zile_dp)
                medii['Total']    = sum(medii.values())
            else:
                # Fallback la calculul clasic dacă Date prelucrate nu are date
                sume = {}
                for i in range(1, 9):
                    sume[f'Clasa_{i}'] = (df_calcul[f'B1_Clasa_{i}'].sum() +
                                           df_calcul[f'B2_Clasa_{i}'].sum())
                sume['Clasa_15'] = (df_calcul['B1_Clasa_15'].sum() +
                                     df_calcul['B2_Clasa_15'].sum())
                medii = {f'Clasa_{i}': round(sume[f'Clasa_{i}'] / divisor) for i in range(1, 9)}
                medii['Clasa_15'] = round(sume['Clasa_15'] / divisor)
                medii['Total']    = sum(medii.values())
        else:
            medii = {f'Clasa_{i}': 0 for i in range(1, 9)}
            medii['Clasa_15'] = 0; medii['Total'] = 0
        an, luna = row['An'], row['Luna']
        df_luna_valida = df_valid[(df_valid['An'] == an) & (df_valid['Luna'] == luna)]
        nr_zile_valide = df_luna_valida['Zi'].nunique() if len(df_luna_valida) > 0 else 0
        zile_in_luna   = calendar.monthrange(an, luna)[1]

        if nr_zile_valide == zile_in_luna:
            indicator, color = "Date complete", "NORMAL"
            df_calcul = df_luna_valida; divisor = zile_in_luna
        elif nr_zile_valide >= MIN_ZILE_LUNA:
            indicator, color = "Date parțiale", "YELLOW"
            df_calcul = df_luna_valida; divisor = nr_zile_valide
        else:
            has_7, week_dates = check_7_consecutive_days(df, an, luna)
            if has_7:
                indicator, color = "Date parțiale - 7 zile", "YELLOW"
                df_calcul = df[df['Zi'].isin(week_dates)]; divisor = 7
            else:
                indicator, color = "Nu există date", "ORANGE"
                df_calcul = pd.DataFrame(); divisor = 0

        if divisor > 0 and not df_calcul.empty:
            sume = {}
            for i in range(1, 9):
                sume[f'Clasa_{i}'] = (df_calcul[f'B1_Clasa_{i}'].sum() +
                                       df_calcul[f'B2_Clasa_{i}'].sum())
            sume['Clasa_15'] = (df_calcul['B1_Clasa_15'].sum() +
                                 df_calcul['B2_Clasa_15'].sum())
            medii = {f'Clasa_{i}': round(sume[f'Clasa_{i}'] / divisor) for i in range(1, 9)}
            medii['Clasa_15'] = round(sume['Clasa_15'] / divisor)
            medii['Total']    = sum(medii.values())
        else:
            medii = {f'Clasa_{i}': 0 for i in range(1, 9)}
            medii['Clasa_15'] = 0; medii['Total'] = 0

        # Mod funcționare lunar = tipul zilei majoritar din Date prelucrate
        # pentru zilele din luna an/luna
        _tip_counts = {'clasificator': 0, 'totalizator': 0, 'null': 0}
        for _zi_d, _tip in _dp_tip_zi_map.items():
            _zi_an  = _zi_d.year if hasattr(_zi_d, 'year') else int(str(_zi_d)[:4])
            _zi_lu  = _zi_d.month if hasattr(_zi_d, 'month') else int(str(_zi_d)[5:7])
            if _zi_an == an and _zi_lu == luna:
                _tip_counts[_tip] = _tip_counts.get(_tip, 0) + 1
        if sum(_tip_counts.values()) > 0:
            mod_functionare_lunar = max(_tip_counts, key=_tip_counts.get)
        else:
            mod_functionare_lunar = 'clasificator'

        lunar_data.append({
            'Post': site_id, 'An': an, 'Luna': luna,
            **medii, 'Indicator': indicator, 'Color': color,
            'Zile_cu_inregistrari': f"{nr_zile_valide}/{zile_in_luna}",
            'Are_date_valide': (divisor > 0),
            'Mod_functionare': mod_functionare_lunar,
        })

    # ── Aplică suprascrierile manuale ─────────────────────────────────────────
    for entry in lunar_data:
        key = (int(entry['An']), int(entry['Luna']))
        if key in _manual_overrides:
            mzl_m = round(_manual_overrides[key])
            # Distribuim egal Total manual pe clase (păstrăm proporțiile dacă există date)
            old_total = entry.get('Total', 0)
            if old_total and old_total > 0:
                ratio = mzl_m / old_total
                for cls in [f'Clasa_{i}' for i in range(1, 9)] + ['Clasa_15']:
                    entry[cls] = round(entry.get(cls, 0) * ratio)
            else:
                # Nu există date calculate — punem tot în Total, clase 0
                for cls in [f'Clasa_{i}' for i in range(1, 9)] + ['Clasa_15']:
                    entry[cls] = 0
            entry['Total']    = mzl_m
            entry['Color']    = 'MANUAL'
            entry['Indicator'] = f"Prelucrare manuală (MZL={mzl_m})"

    # headers_lunar definit ÎNAINTE de merge_cells care îl folosește
    headers_lunar = ["Post","An","Luna","Clasa 1","Clasa 2","Clasa 3","Clasa 4",
                     "Clasa 5","Clasa 6","Clasa 7","Clasa 8","Clasa 15",
                     "Total","Indicator","Mod de funcționare","Zile cu înregistrări"]

    ws_lunar.merge_cells(f"A1:{get_column_letter(len(headers_lunar))}1")
    ws_lunar["A1"] = (f"MEDIE ZILNICĂ LUNARĂ  |  Contor: {site_id}"
                      + (f"  |  {localitate_site.lstrip(' -')}" if localitate_site else "")
                      + f"  |  Minim {MIN_ORE_ZI}h/zi, minim {MIN_ZILE_LUNA} zile/lună")
    ws_lunar["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws_lunar["A1"].fill = fill(C_DARK)
    ws_lunar["A1"].alignment = ctr()
    ws_lunar.row_dimensions[1].height = 28

    for c, h in enumerate(headers_lunar, 1):
        cell = ws_lunar.cell(2, c, h)
        cell.font = hfont(9); cell.fill = fill(C_MID)
        cell.alignment = ctr(); cell.border = brd
    ws_lunar.row_dimensions[2].height = 28

    luna_nume = {1:'Ian',2:'Feb',3:'Mar',4:'Apr',5:'Mai',6:'Iun',
                 7:'Iul',8:'Aug',9:'Sep',10:'Oct',11:'Noi',12:'Dec'}

    dr_lunar = 3
    for i, rd in enumerate(lunar_data):
        if rd['Color'] == 'MANUAL':
            rf = fill(C_VIOLET)
        elif rd['Color'] == 'YELLOW':
            rf = fill(C_YELLOW)
        elif rd['Color'] == 'ORANGE':
            rf = fill(C_ORANGE)
        else:
            rf = fill(C_LIGHT) if i % 2 == 0 else fill(C_WHITE)

        vals = [rd['Post'], rd['An'], luna_nume[rd['Luna']],
                rd['Clasa_1'], rd['Clasa_2'], rd['Clasa_3'], rd['Clasa_4'],
                rd['Clasa_5'], rd['Clasa_6'], rd['Clasa_7'], rd['Clasa_8'],
                rd['Clasa_15'], rd['Total'], rd['Indicator'],
                rd.get('Mod_functionare', '').capitalize(),
                rd['Zile_cu_inregistrari']]
        for c, val in enumerate(vals, 1):
            cell = ws_lunar.cell(dr_lunar, c, val)
            cell.font = dfont(9); cell.fill = rf; cell.border = brd
            # cols 1-3=text centrat, 4-13=numere dreapta, 14+=text centrat
            cell.alignment = ctr() if c <= 3 or c >= 14 else rgt()
            if 4 <= c <= 13:
                cell.number_format = '#,##0'
        dr_lunar += 1

    # ══════════════════════════════════════════════════════════════════════════
    # MZA — Media Zilnică Anuală (sub tabelul MZL, per an)
    # ══════════════════════════════════════════════════════════════════════════
    # Reguli:
    #   Lună validă = Are_date_valide True (≥ MIN_ZILE_LUNA zile sau 7 consecutive)
    #   MZA = suma(MZL_luni_valide) / nr_luni_valide, dacă ≥ _MIN_LUNI_AN luni valide
    #   Fallback 1: dacă < _MIN_LUNI_AN luni valide → valoarea lunii Mai (_MIN_LUNI_AN_MAI)
    #   Fallback 2: dacă Mai lipsește → valoarea lunii Octombrie (_MIN_LUNI_AN_OCT)
    #   Fără date: celulă goală

    C_MZA_HDR  = "1F4E79"   # header MZA — albastru închis
    C_MZA_NORM = "DEEAF1"   # rând MZA calculat normal
    C_MZA_FALL = "FCE4D6"   # rând MZA fallback (portocaliu pal)
    C_MZA_NULL = "F4CCCC"   # rând MZA fără date

    lunar_df  = pd.DataFrame(lunar_data)
    ani_unici = sorted(lunar_df['An'].unique().astype(int).tolist())

    # Spațiu separator
    dr_sep = dr_lunar + 1

    # Titlu secțiune MZA
    n_cols_lunar = len(headers_lunar)
    ws_lunar.merge_cells(f"A{dr_sep}:{get_column_letter(n_cols_lunar)}{dr_sep}")
    _mza_title = ws_lunar[f"A{dr_sep}"]
    _mza_title.value     = (f"MEDIA ZILNICĂ ANUALĂ (MZA)  |  "
                             f"Minim {_MIN_LUNI_AN} luni valide/an  |  "
                             f"Fallback: Mai → Octombrie")
    _mza_title.font      = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    _mza_title.fill      = fill(C_MZA_HDR)
    _mza_title.alignment = ctr()
    ws_lunar.row_dimensions[dr_sep].height = 24
    dr_sep += 1

    # Header MZA — primele 3 coloane diferite față de MZL, ultima coloană = "Luni cu înregistrări"
    _mza_hdrs = ["Contor", "An", "Tip"] + headers_lunar[3:-1] + ["Luni cu înregistrări"]
    for c_i, h in enumerate(_mza_hdrs, 1):
        _hc = ws_lunar.cell(dr_sep, c_i, h)
        _hc.font      = Font(name='Arial', size=9, bold=True, color="FFFFFF")
        _hc.fill      = fill(C_MZA_HDR)
        _hc.alignment = ctr()
        _hc.border    = brd
    ws_lunar.row_dimensions[dr_sep].height = 22
    dr_sep += 1

    # Calculăm MZA per an
    for an_mza in ani_unici:
        df_an = lunar_df[lunar_df['An'] == an_mza]

        # Luni valide = Are_date_valide == True
        df_valide = df_an[df_an['Are_date_valide'] == True]
        n_luni_val = len(df_valide)

        clase_cols_mza = [f'Clasa_{i}' for i in range(1, 9)] + ['Clasa_15']

        if n_luni_val >= _MIN_LUNI_AN:
            # Calcul normal: media pe luni valide
            medii_mza = {}
            for col in clase_cols_mza:
                if col in df_valide.columns:
                    medii_mza[col] = round(df_valide[col].sum() / n_luni_val)
                else:
                    medii_mza[col] = 0
            medii_mza['Total'] = sum(medii_mza.values())
            indicator_mza = f"MZA normală ({n_luni_val} luni)"
            color_mza     = C_MZA_NORM
            mod_mza       = "Clasificator"

        else:
            # Fallback 1: luna Mai
            df_mai = df_an[df_an['Luna'] == _MIN_LUNI_AN_MAI]
            df_mai_valid = df_mai[df_mai['Are_date_valide'] == True]

            if not df_mai_valid.empty:
                medii_mza = {}
                for col in clase_cols_mza:
                    medii_mza[col] = int(df_mai_valid.iloc[0].get(col, 0))
                medii_mza['Total'] = sum(medii_mza.values())
                indicator_mza = f"Fallback Mai ({n_luni_val} luni valide)"
                color_mza     = C_MZA_FALL
                mod_mza       = df_mai_valid.iloc[0].get('Mod_functionare', 'clasificator').capitalize()
            else:
                # Fallback 2: luna Octombrie
                df_oct = df_an[df_an['Luna'] == _MIN_LUNI_AN_OCT]
                df_oct_valid = df_oct[df_oct['Are_date_valide'] == True]

                if not df_oct_valid.empty:
                    medii_mza = {}
                    for col in clase_cols_mza:
                        medii_mza[col] = int(df_oct_valid.iloc[0].get(col, 0))
                    medii_mza['Total'] = sum(medii_mza.values())
                    indicator_mza = f"Fallback Octombrie ({n_luni_val} luni valide)"
                    color_mza     = C_MZA_FALL
                    mod_mza       = df_oct_valid.iloc[0].get('Mod_functionare', 'clasificator').capitalize()
                else:
                    # Fără date suficiente
                    medii_mza     = {col: 0 for col in clase_cols_mza}
                    medii_mza['Total'] = 0
                    indicator_mza = f"Insuficient ({n_luni_val} luni valide)"
                    color_mza     = C_MZA_NULL
                    mod_mza       = "—"

        rf_mza = fill(color_mza)
        vals_mza = [
            site_id, an_mza, "MZA",
            medii_mza.get('Clasa_1',0), medii_mza.get('Clasa_2',0),
            medii_mza.get('Clasa_3',0), medii_mza.get('Clasa_4',0),
            medii_mza.get('Clasa_5',0), medii_mza.get('Clasa_6',0),
            medii_mza.get('Clasa_7',0), medii_mza.get('Clasa_8',0),
            medii_mza.get('Clasa_15',0), medii_mza.get('Total',0),
            indicator_mza, mod_mza, f"{n_luni_val}/12 luni",
        ]
        for c_i, val in enumerate(vals_mza, 1):
            _c = ws_lunar.cell(dr_sep, c_i, val)
            _c.font      = dfont(9, bold=True)
            _c.fill      = rf_mza
            _c.border    = brd
            _c.alignment = ctr() if c_i <= 3 or c_i >= 14 else rgt()
            if 4 <= c_i <= 13 and isinstance(val, int):
                _c.number_format = '#,##0'
        ws_lunar.row_dimensions[dr_sep].height = 22
        dr_sep += 1

    # Lățimi coloane — actualizate pentru noile coloane
    widths_lunar = [10, 8, 8] + [10] * 9 + [12, 20, 18, 16]
    for c, w in enumerate(widths_lunar, 1):
        ws_lunar.column_dimensions[get_column_letter(c)].width = w

    # Grafic lunar
    lunar_df  = pd.DataFrame(lunar_data)
    ani_unici = sorted(lunar_df['An'].unique())
    # start_chart_row trebuie să fie după tabelul MZA (dr_sep) + 2 rânduri separator
    start_chart_row = dr_sep + 2

    ws_lunar.cell(start_chart_row, 1, "Luna").font = Font(bold=True)
    for idx_an, an in enumerate(ani_unici, 2):
        cell = ws_lunar.cell(start_chart_row, idx_an, an)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    luna_nome_full = ['Ianuarie','Februarie','Martie','Aprilie','Mai','Iunie',
                      'Iulie','August','Septembrie','Octombrie','Noiembrie','Decembrie']
    for luna_nr in range(1, 13):
        r_idx = start_chart_row + luna_nr
        ws_lunar.cell(r_idx, 1, luna_nome_full[luna_nr - 1])
        for idx_an, an in enumerate(ani_unici, 2):
            row_v   = lunar_df[(lunar_df['An'] == an) & (lunar_df['Luna'] == luna_nr)]
            valoare = int(row_v.iloc[0]['Total']) if len(row_v) > 0 else 0
            ws_lunar.cell(r_idx, idx_an, valoare).number_format = '#,##0'

    cats_l = Reference(ws_lunar, min_col=1, min_row=start_chart_row+1, max_row=start_chart_row+12)
    data_l = Reference(ws_lunar, min_col=2, max_col=1+len(ani_unici),
                       min_row=start_chart_row, max_row=start_chart_row+12)

    chart_lunar = LineChart()
    chart_lunar.title  = "Media Zilnică Lunară - Total vehicule"
    chart_lunar.style  = 10; chart_lunar.width = 25; chart_lunar.height = 14
    chart_lunar.add_data(data_l, titles_from_data=True, from_rows=False)
    chart_lunar.set_categories(cats_l)
    chart_lunar.x_axis.delete = False; chart_lunar.y_axis.delete = False
    chart_lunar.legend.position = 'b'; chart_lunar.legend.overlay = False
    chart_lunar.x_axis.axPos = "b"; chart_lunar.x_axis.tickLblPos = "nextTo"
    chart_lunar.x_axis.textRotation = -45000
    chart_lunar.y_axis.scaling.min = 0
    chart_lunar.y_axis.title = "Număr vehicule"
    chart_lunar.layout = Layout(manualLayout=ManualLayout(x=0.1, y=0.12, h=0.7, w=0.85,
                                                          xMode="edge", yMode="edge"))
    culori_ani = ['2E75B6','ED7D31','70AD47','FFC000','5B9BD5']
    for idx, serie in enumerate(chart_lunar.series):
        serie.graphicalProperties.line.solidFill = culori_ani[idx % len(culori_ani)]
        serie.graphicalProperties.line.width = 25000
        serie.smooth = True; serie.marker.symbol = "circle"; serie.marker.size = 7
        serie.marker.graphicalProperties.solidFill = culori_ani[idx % len(culori_ani)]
        serie.marker.graphicalProperties.line.solidFill = culori_ani[idx % len(culori_ani)]
    ws_lunar.add_chart(chart_lunar, "Q3")

    # ── FOAIE 4: Profil Orar Mediu ───────────────────────────────────────────
    ws3 = wb.create_sheet("Profil Orar Mediu")

    # Coloane: Ora | Media_B1..BN | Media Total | B1%..BN%
    # 2 benzi: 1+2+1+2 = 6 col
    # 4 benzi: 1+4+1+4 = 10 col
    # 6 benzi: 1+6+1+6 = 14 col
    n_cols_ws3 = 1 + n_bands + 1 + n_bands  # Ora + medii + total + procente
    ws3_title_end = get_column_letter(n_cols_ws3)
    ws3.merge_cells(f"A1:{ws3_title_end}1")
    ws3["A1"] = f"PROFIL ORAR MEDIU (toate zilele)  |  Contor: {site_id}"
    ws3["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws3["A1"].fill = fill(C_DARK); ws3["A1"].alignment = ctr()
    ws3.row_dimensions[1].height = 28

    h3 = ["Ora"] + [f"Media Banda {b}" for b in band_ids] + ["Media Total"] + \
         [f"Banda {b} %" for b in band_ids]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(2, c, h)
        cell.font = hfont(); cell.fill = fill(C_MID)
        cell.alignment = ctr(); cell.border = brd
    ws3.row_dimensions[2].height = 28

    # Coloane: Ora=1, Media_B1=2..n_bands+1, Total=n_bands+2, Proc_B1=n_bands+3..
    media_start = 2
    total_col_ws3 = media_start + n_bands  # ex: 4 benzi -> col 6
    proc_start = total_col_ws3 + 1

    # Formula total = SUM(B..E) la ora respectivă
    for row_i in range(24):
        dr3    = row_i + 3
        rf     = fill(C_LIGHT) if row_i % 2 == 0 else fill(C_WHITE)

        vals = [f"{row_i:02d}:00"]
        for b in band_ids:
            hr_data = hourly_avg[hourly_avg['Ora'] == row_i]
            avg_b = int(round(hr_data[f'Total_B{b}'].values[0])) if len(hr_data) > 0 else 0
            vals.append(avg_b)

        # Media Total = formula SUM
        band_cols_letters = [get_column_letter(media_start + bi) for bi in range(n_bands)]
        total_formula = "=" + "+".join([f"{cl}{dr3}" for cl in band_cols_letters])
        vals.append(total_formula)

        total_cl = get_column_letter(total_col_ws3)
        for bi in range(n_bands):
            b_cl = get_column_letter(media_start + bi)
            vals.append(f"=IFERROR({b_cl}{dr3}/{total_cl}{dr3},0)")

        for c, val in enumerate(vals, 1):
            cell = ws3.cell(dr3, c, val)
            cell.font = dfont(); cell.fill = rf; cell.border = brd
            cell.alignment = ctr() if c == 1 else rgt()
            if 2 <= c <= total_col_ws3: cell.number_format = '#,##0'
            if c >= proc_start:         cell.number_format = '0.0%'

    # Ora 50
    ani_pentru_ora50 = sorted(df['An'].unique())
    dr_ora50_start   = 28
    ws3.cell(dr_ora50_start - 1, 1, "Top Trafic - Ora 50 (pe an)").font = Font(bold=True)
    thin2 = Side(style='thin', color='BFBFBF')
    brd2  = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

    for idx, an_curent in enumerate(ani_pentru_ora50):
        df_an = df[df['An'] == an_curent].copy()
        if len(df_an) < 50: continue
        df_sortat  = df_an.sort_values(by='Total_General', ascending=False).reset_index(drop=True)
        ora_50_row = df_sortat.iloc[49]
        ora_nr     = int(pd.to_datetime(ora_50_row['Data']).hour)
        ora_data_str = pd.to_datetime(ora_50_row['Data']).strftime('%d.%m.%Y %H:00')
        ora_total  = int(ora_50_row['Total_General'])
        rind_curent = dr_ora50_start + idx

        cell_desc = ws3.cell(rind_curent, 1)
        cell_desc.value = f"Ora 50 anul {an_curent}\n({ora_nr:02d}:00 - {ora_data_str})"
        cell_desc.font  = Font(name='Arial', size=9, bold=True, color="FFFFFF")
        cell_desc.fill  = PatternFill('solid', start_color="ED7D31")
        cell_desc.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
        cell_desc.border = brd2
        ws3.row_dimensions[rind_curent].height = 45

        # Valorile pe fiecare bandă + total + procente
        ora_vals_banda = []
        for b in band_ids:
            ora_vals_banda.append(int(ora_50_row.get(f'Total_B{b}', 0)))
        ora_vals_banda.append(ora_total)
        for b_idx in range(n_bands):
            v = ora_vals_banda[b_idx]
            ora_vals_banda.append(v / ora_total if ora_total > 0 else 0)

        for col_idx, valoare in enumerate(ora_vals_banda, 2):
            cell = ws3.cell(rind_curent, col_idx, valoare)
            cell.font = Font(name='Arial', size=9, bold=True)
            cell.border = brd2
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if col_idx == total_col_ws3:  # col Total
                cell.fill = PatternFill('solid', start_color="FCE4D6")
                cell.font = Font(name='Arial', size=9, bold=True, color="C65911")
            is_proc_col = col_idx >= proc_start
            cell.number_format = '0.0%' if is_proc_col else '#,##0'

    # Lățimi coloane ws3
    col_widths_ws3 = [20] + [16] * n_bands + [14] + [12] * n_bands
    for c, w in enumerate(col_widths_ws3, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # Grafic 3: profil orar line — pe SENSURI (sens1 vs sens2)
    chart3 = LineChart()
    sens1_label = " + ".join([f"B{b}" for b in sens1_bands])
    sens2_label = " + ".join([f"B{b}" for b in sens2_bands])
    chart3.title = f"Profil orar mediu — Sens 1 ({sens1_label}) vs Sens 2 ({sens2_label})"
    chart3.y_axis.title = "Vehicule / oră (medie)"; chart3.x_axis.title = "Ora"
    chart3.style = 10; chart3.width = 20; chart3.height = 14
    chart3.legend.position = 'b'; chart3.legend.overlay = False
    chart3.x_axis.delete = False; chart3.y_axis.delete = False
    chart3.x_axis.axPos = "b"; chart3.x_axis.tickLblPos = "nextTo"
    chart3.x_axis.textRotation = -45000; chart3.x_axis.tickLblSkip = 1
    chart3.layout = Layout(manualLayout=ManualLayout(x=0.1,y=0.12,h=0.7,w=0.85,
                                                     xMode="edge",yMode="edge"))

    # Plasăm date auxiliare pentru grafic sensuri în coloane extra ale ws3
    # IMPORTANT: calculăm valorile sens ca suma benzilor individuale deja rotunjite,
    # astfel s1 + s2 = exact suma tuturor benzilor (fără eroare de rotunjire)
    chart3_data_col_start = n_cols_ws3 + 2
    ws3.cell(2, chart3_data_col_start, "Sens 1").font = dfont(bold=True)
    ws3.cell(2, chart3_data_col_start + 1, "Sens 2").font = dfont(bold=True)
    for row_i in range(24):
        dr3 = row_i + 3
        hr_data = hourly_avg[hourly_avg['Ora'] == row_i]
        # Calculăm sens ca suma benzilor individuale rotunjite (nu din Total_Sens1/2)
        s1_val = sum(
            int(round(hr_data[f'Total_B{b}'].values[0])) if len(hr_data) > 0 else 0
            for b in sens1_bands
        )
        s2_val = sum(
            int(round(hr_data[f'Total_B{b}'].values[0])) if len(hr_data) > 0 else 0
            for b in sens2_bands
        )
        ws3.cell(dr3, chart3_data_col_start, s1_val)
        ws3.cell(dr3, chart3_data_col_start + 1, s2_val)

    ref_s1  = Reference(ws3, min_col=chart3_data_col_start,     min_row=2, max_row=26)
    ref_s2  = Reference(ws3, min_col=chart3_data_col_start + 1, min_row=2, max_row=26)
    cats3   = Reference(ws3, min_col=1, min_row=3, max_row=26)
    chart3.add_data(ref_s1, titles_from_data=True)
    chart3.add_data(ref_s2, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chart3.series[0].graphicalProperties.line.width = 20000
    chart3.series[0].marker.symbol = "circle"; chart3.series[0].marker.size = 5
    chart3.series[1].graphicalProperties.line.solidFill = "ED7D31"
    chart3.series[1].graphicalProperties.line.width = 20000
    chart3.series[1].marker.symbol = "circle"; chart3.series[1].marker.size = 5

    chart3_anchor_col = get_column_letter(n_cols_ws3 + 5)
    ws3.add_chart(chart3, f"{chart3_anchor_col}2")

    # Grafic 4: stacked procentual orar — pe BENZI individuale
    chart4 = BarChart()
    chart4.type = "col"; chart4.grouping = "percentStacked"; chart4.overlap = 100
    chart4.title = f"Distribuție procentuală orară — {'  /  '.join([f'Banda {b}' for b in band_ids])}"
    chart4.y_axis.title = "Procent (%)"; chart4.x_axis.title = "Ora"
    chart4.style = 10; chart4.width = 20; chart4.height = 14
    chart4.legend.position = 'b'; chart4.legend.overlay = False
    chart4.x_axis.delete = False; chart4.x_axis.axPos = "b"
    chart4.x_axis.tickLblPos = "nextTo"; chart4.x_axis.textRotation = -45000
    chart4.x_axis.tickLblSkip = 1
    chart4.layout = Layout(manualLayout=ManualLayout(x=0.1,y=0.12,h=0.7,w=0.85,
                                                     xMode="edge",yMode="edge"))

    for bi in range(n_bands):
        col_idx = media_start + bi
        ref_b = Reference(ws3, min_col=col_idx, min_row=2, max_row=26)
        chart4.add_data(ref_b, titles_from_data=True)
        chart4.series[bi].graphicalProperties.solidFill = BAND_COLORS[bi % len(BAND_COLORS)]
    chart4.set_categories(cats3)

    chart4_anchor_col = get_column_letter(n_cols_ws3 + 5)
    ws3.add_chart(chart4, f"{chart4_anchor_col}30")

    # ── FOAIE 5: Profil Zilnic Mediu ─────────────────────────────────────────
    ws3z = wb.create_sheet("Profil Zilnic Mediu")

    # Structura identica cu Profil Orar Mediu, dar pe 7 zile de saptamana
    # Coloane: Zi | Media_B1..BN | Media Total | B1%..BN%
    n_cols_ws3z = 1 + n_bands + 1 + n_bands
    ws3z_title_end = get_column_letter(n_cols_ws3z)
    ws3z.merge_cells(f"A1:{ws3z_title_end}1")
    ws3z["A1"] = f"PROFIL ZILNIC MEDIU (toate săptămânile)  |  Contor: {site_id}"
    ws3z["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws3z["A1"].fill = fill(C_DARK); ws3z["A1"].alignment = ctr()
    ws3z.row_dimensions[1].height = 28

    h3z = ["Zi"] + [f"Media Banda {b}" for b in band_ids] + ["Media Total"] + \
          [f"Banda {b} %" for b in band_ids]
    for c, h in enumerate(h3z, 1):
        cell = ws3z.cell(2, c, h)
        cell.font = hfont(); cell.fill = fill(C_MID)
        cell.alignment = ctr(); cell.border = brd
    ws3z.row_dimensions[2].height = 28

    total_col_ws3z = 2 + n_bands   # col Media Total
    proc_start_z   = total_col_ws3z + 1

    # Calculam media pe zi de saptamana din daily_data
    # daily_data are col 'Zi' (date) si Total_Bx
    # Adaugam coloana weekday (0=Luni..6=Duminica)
    dd_z = daily_data.copy()
    dd_z['weekday'] = pd.to_datetime(dd_z['Zi']).dt.weekday  # 0=Luni

    # Media per banda per zi de saptamana (doar zilele cu >= MIN_ORE_ZI ore)
    # Folosim df direct, care are 'Ora' si 'Zi'
    ore_pe_zi_df = df.groupby('Zi').size().reset_index(name='Ore_pe_zi')
    dd_z = dd_z.merge(ore_pe_zi_df, on='Zi', how='left')
    dd_z_valid = dd_z[dd_z['Ore_pe_zi'] >= MIN_ORE_ZI]

    weekly_avg = dd_z_valid.groupby('weekday').agg(
        {f'Total_B{b}': 'mean' for b in band_ids}
    ).reset_index()

    for row_i, day_name in enumerate(days_ro):  # 0=Luni..6=Duminica
        dr3z = row_i + 3
        rf   = fill(C_LIGHT) if row_i % 2 == 0 else fill(C_WHITE)

        vals = [day_name]
        for b in band_ids:
            wd_data = weekly_avg[weekly_avg['weekday'] == row_i]
            avg_b = int(round(wd_data[f'Total_B{b}'].values[0])) \
                    if len(wd_data) > 0 else 0
            vals.append(avg_b)

        # Media Total = formula SUM
        band_cols_z = [get_column_letter(2 + bi) for bi in range(n_bands)]
        total_formula_z = "=" + "+".join([f"{cl}{dr3z}" for cl in band_cols_z])
        vals.append(total_formula_z)

        total_cl_z = get_column_letter(total_col_ws3z)
        for bi in range(n_bands):
            b_cl = get_column_letter(2 + bi)
            vals.append(f"=IFERROR({b_cl}{dr3z}/{total_cl_z}{dr3z},0)")

        for c, val in enumerate(vals, 1):
            cell = ws3z.cell(dr3z, c, val)
            cell.font = dfont(); cell.fill = rf; cell.border = brd
            cell.alignment = ctr() if c == 1 else rgt()
            if 2 <= c <= total_col_ws3z: cell.number_format = '#,##0'
            if c >= proc_start_z:        cell.number_format = '0.0%'

    # Latimi coloane
    col_widths_z = [16] + [16] * n_bands + [14] + [12] * n_bands
    for c, w in enumerate(col_widths_z, 1):
        ws3z.column_dimensions[get_column_letter(c)].width = w

    # ── Grafic Z1: profil zilnic line — Sens 1 vs Sens 2 ─────────────────────
    chartZ1 = LineChart()
    chartZ1.title = f"Profil zilnic mediu — Sens 1 ({sens1_label}) vs Sens 2 ({sens2_label})"
    chartZ1.y_axis.title = "Vehicule / zi (medie)"
    chartZ1.x_axis.title = "Ziua săptămânii"
    chartZ1.style = 10; chartZ1.width = 20; chartZ1.height = 14
    chartZ1.legend.position = 'b'; chartZ1.legend.overlay = False
    chartZ1.x_axis.delete = False; chartZ1.y_axis.delete = False
    chartZ1.x_axis.axPos = "b"; chartZ1.x_axis.tickLblPos = "nextTo"
    chartZ1.layout = Layout(manualLayout=ManualLayout(
        x=0.1, y=0.12, h=0.7, w=0.85, xMode="edge", yMode="edge"))

    # Date auxiliare pentru sensuri in coloane extra
    cz_data_col = n_cols_ws3z + 2
    ws3z.cell(2, cz_data_col,     "Sens 1").font = dfont(bold=True)
    ws3z.cell(2, cz_data_col + 1, "Sens 2").font = dfont(bold=True)
    for row_i in range(7):
        drz = row_i + 3
        wd_data = weekly_avg[weekly_avg['weekday'] == row_i]
        s1_val = sum(
            int(round(wd_data[f'Total_B{b}'].values[0])) if len(wd_data) > 0 else 0
            for b in sens1_bands
        )
        s2_val = sum(
            int(round(wd_data[f'Total_B{b}'].values[0])) if len(wd_data) > 0 else 0
            for b in sens2_bands
        )
        ws3z.cell(drz, cz_data_col,     s1_val)
        ws3z.cell(drz, cz_data_col + 1, s2_val)

    ref_z_s1 = Reference(ws3z, min_col=cz_data_col,     min_row=2, max_row=9)
    ref_z_s2 = Reference(ws3z, min_col=cz_data_col + 1, min_row=2, max_row=9)
    cats_z   = Reference(ws3z, min_col=1, min_row=3, max_row=9)
    chartZ1.add_data(ref_z_s1, titles_from_data=True)
    chartZ1.add_data(ref_z_s2, titles_from_data=True)
    chartZ1.set_categories(cats_z)
    chartZ1.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chartZ1.series[0].graphicalProperties.line.width = 20000
    chartZ1.series[0].marker.symbol = "circle"; chartZ1.series[0].marker.size = 6
    chartZ1.series[1].graphicalProperties.line.solidFill = "ED7D31"
    chartZ1.series[1].graphicalProperties.line.width = 20000
    chartZ1.series[1].marker.symbol = "circle"; chartZ1.series[1].marker.size = 6

    cz_anchor = get_column_letter(n_cols_ws3z + 5)
    ws3z.add_chart(chartZ1, f"{cz_anchor}2")

    # ── Grafic Z2: stacked procentual zilnic — benzi individuale ──────────────
    chartZ2 = BarChart()
    chartZ2.type = "col"; chartZ2.grouping = "percentStacked"; chartZ2.overlap = 100
    chartZ2.title = f"Distribuție procentuală zilnică — " \
                    f"{'  /  '.join([f'Banda {b}' for b in band_ids])}"
    chartZ2.y_axis.title = "Procent (%)"
    chartZ2.x_axis.title = "Ziua săptămânii"
    chartZ2.style = 10; chartZ2.width = 20; chartZ2.height = 14
    chartZ2.legend.position = 'b'; chartZ2.legend.overlay = False
    chartZ2.x_axis.delete = False; chartZ2.x_axis.axPos = "b"
    chartZ2.x_axis.tickLblPos = "nextTo"
    chartZ2.layout = Layout(manualLayout=ManualLayout(
        x=0.1, y=0.12, h=0.7, w=0.85, xMode="edge", yMode="edge"))

    for bi in range(n_bands):
        col_idx_z = 2 + bi
        ref_bz = Reference(ws3z, min_col=col_idx_z, min_row=2, max_row=9)
        chartZ2.add_data(ref_bz, titles_from_data=True)
        chartZ2.series[bi].graphicalProperties.solidFill = BAND_COLORS[bi % len(BAND_COLORS)]
    chartZ2.set_categories(cats_z)

    ws3z.add_chart(chartZ2, f"{cz_anchor}30")

    # ── FOAIE 6: Comparație Benzi ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Comparatie Benzi")

    sens1_lbl = " + ".join([f"Banda {b}" for b in sens1_bands])
    sens2_lbl = " + ".join([f"Banda {b}" for b in sens2_bands])

    n_cols_ws4 = 2 + n_bands + 2 + 1  # Indicator + benzi + Diferenta + Raport
    ws4_title_end = get_column_letter(n_cols_ws4)
    ws4.merge_cells(f"A1:{ws4_title_end}1")
    ws4["A1"] = f"COMPARAȚIE BENZI  |  Contor: {site_id}"
    ws4["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws4["A1"].fill = fill(C_DARK); ws4["A1"].alignment = ctr()
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells(f"A2:{ws4_title_end}2")
    ws4["A2"] = f"Sens 1: {sens1_lbl}  |  Sens 2: {sens2_lbl}"
    ws4["A2"].font = dfont(9, color="595959"); ws4["A2"].alignment = ctr()
    ws4.row_dimensions[2].height = 20

    headers4 = ["Indicator"] + [f"Banda {b}" for b in band_ids] + \
               [f"Diferență (Sens1–Sens2)", f"Raport Sens1/Sens2"]
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(3, c, h)
        cell.font = hfont(); cell.fill = fill(C_MID)
        cell.alignment = ctr(); cell.border = brd
    ws4.row_dimensions[3].height = 28

    totals_per_band = {b: int(df[f'Total_B{b}'].sum()) for b in band_ids}
    peaks_per_band  = {b: int(df[f'Total_B{b}'].max()) for b in band_ids}
    avg_daily_band  = {b: round(totals_per_band[b] / n_days) for b in band_ids}
    avg_hourly_band = {b: round(df[f'Total_B{b}'].mean()) for b in band_ids}

    total_sens1 = sum(totals_per_band[b] for b in sens1_bands)
    total_sens2 = sum(totals_per_band[b] for b in sens2_bands)
    total_all   = total_sens1 + total_sens2

    def sens_diff_ratio(metric_dict):
        s1 = sum(metric_dict[b] for b in sens1_bands)
        s2 = sum(metric_dict[b] for b in sens2_bands)
        return s1 - s2, (s1 / s2 if s2 else 0)

    rows4_data = [
        ("Total vehicule (toată perioada)", totals_per_band),
        ("Medie zilnică (veh/zi)", avg_daily_band),
        ("Medie orară (veh/h)", avg_hourly_band),
        ("Vârf maxim înregistrat (veh/h)", peaks_per_band),
        ("Procent din trafic total", None),
    ]

    dr4 = 4
    for i, (label, data_dict) in enumerate(rows4_data):
        rf = fill(C_LIGHT) if i % 2 == 0 else fill(C_WHITE)
        if label == "Procent din trafic total":
            row_vals = [label]
            for b in band_ids:
                row_vals.append(f"={totals_per_band[b]}/({total_all})")
            row_vals += ["—", "—"]
            fmts = [None] + ['0.0%'] * n_bands + [None, None]
        else:
            diff, ratio = sens_diff_ratio(data_dict)
            row_vals = [label] + [data_dict[b] for b in band_ids] + [diff, f"{ratio:.3f}x"]
            fmts = [None] + ['#,##0'] * n_bands + ['#,##0', None]

        for c, (val, fmt) in enumerate(zip(row_vals, fmts), 1):
            cell = ws4.cell(dr4, c, val)
            cell.font = dfont(bold=(c==1)); cell.fill = rf
            cell.border = brd; cell.alignment = ctr() if c==1 else rgt()
            if fmt: cell.number_format = fmt
            diff_col = 1 + n_bands + 1
            if c == diff_col and isinstance(val, (int, float)):
                cell.font = dfont(bold=True, color=("2E75B6" if val > 0 else "C00000"))
        dr4 += 1

    ws4.column_dimensions['A'].width = 30
    for bi in range(n_bands):
        ws4.column_dimensions[get_column_letter(2 + bi)].width = 15
    ws4.column_dimensions[get_column_letter(2 + n_bands)].width = 20
    ws4.column_dimensions[get_column_letter(3 + n_bands)].width = 16

    # Pie chart — distribuție totală pe benzi
    pie_data_start_col = n_cols_ws4 + 2
    for bi, b in enumerate(band_ids):
        ws4.cell(10 + bi, pie_data_start_col,     f"Banda {b}")
        ws4.cell(10 + bi, pie_data_start_col + 1, totals_per_band[b])

    pie_ws4 = PieChart()
    pct_list = [round(totals_per_band[b] / total_all * 100) for b in band_ids] if total_all else [0]*n_bands
    pie_ws4.title = "Distribuție totală: " + "  |  ".join([f"B{b}({pct_list[bi]}%)" for bi, b in enumerate(band_ids)])
    pie_ws4.style = 10; pie_ws4.width = 14; pie_ws4.height = 12
    pie_data_ref = Reference(ws4, min_col=pie_data_start_col + 1, min_row=10, max_row=10 + n_bands - 1)
    pie_cats_ref = Reference(ws4, min_col=pie_data_start_col,     min_row=10, max_row=10 + n_bands - 1)
    pie_ws4.add_data(pie_data_ref); pie_ws4.set_categories(pie_cats_ref)
    for bi in range(n_bands):
        pt = DataPoint(idx=bi)
        pt.graphicalProperties.solidFill = BAND_COLORS[bi % len(BAND_COLORS)]
        pie_ws4.series[0].dPt.append(pt)
    lbls_ws4 = DataLabelList(); lbls_ws4.showPercent = True; lbls_ws4.showCatName = True
    lbls_ws4.showVal = False; pie_ws4.dLbls = lbls_ws4
    ws4.add_chart(pie_ws4, "J2")

    # Chart trafic zilnic — pe sensuri
    chart_data_start = dr4 + 2
    ws4.cell(chart_data_start, 1, "Data").font = dfont(bold=True)
    ws4.cell(chart_data_start, 2, f"Sens 1 ({sens1_lbl})").font = dfont(bold=True)
    ws4.cell(chart_data_start, 3, f"Sens 2 ({sens2_lbl})").font = dfont(bold=True)
    for ri, row in daily_data.iterrows():
        dr_ch = chart_data_start + 1 + ri
        ws4.cell(dr_ch, 1, row['Zi'].strftime("%d.%m"))
        ws4.cell(dr_ch, 2, int(row['Total_Sens1']))
        ws4.cell(dr_ch, 3, int(row['Total_Sens2']))
    last_chart_row4 = chart_data_start + len(daily_data)

    chart6 = BarChart()
    chart6.type = "col"; chart6.grouping = "clustered"
    chart6.title = f"Trafic zilnic — Sens 1 ({sens1_lbl}) vs Sens 2 ({sens2_lbl})"
    chart6.y_axis.title = "Vehicule / zi"; chart6.style = 10
    chart6.width = latime_dinamica; chart6.height = 12
    chart6.legend.position = 'b'; chart6.legend.overlay = False
    configure_x_axis(chart6.x_axis)
    chart6.y_axis.delete = False
    chart6.layout = make_bar_layout()
    d6s1 = Reference(ws4, min_col=2, min_row=chart_data_start, max_row=last_chart_row4)
    d6s2 = Reference(ws4, min_col=3, min_row=chart_data_start, max_row=last_chart_row4)
    c6ct = Reference(ws4, min_col=1, min_row=chart_data_start+1, max_row=last_chart_row4)
    chart6.add_data(d6s1, titles_from_data=True)
    chart6.add_data(d6s2, titles_from_data=True)
    chart6.set_categories(c6ct)
    chart6.series[0].graphicalProperties.solidFill = "2E75B6"
    chart6.series[1].graphicalProperties.solidFill = "ED7D31"
    ws4.add_chart(chart6, "J26")

    # ── FOAIE 6: Reguli Calcul ───────────────────────────────────────────────
    ws_reguli = wb.create_sheet("Reguli Calcul")
    ws_reguli.merge_cells("A1:D1")
    ws_reguli["A1"] = f"REGULI DE CALCUL  |  Contor: {site_id}"
    ws_reguli["A1"].font = Font(name='Arial', size=13, bold=True, color=C_WHITE)
    ws_reguli["A1"].fill = fill(C_DARK); ws_reguli["A1"].alignment = ctr()
    headers_reguli = ["Regulă","Valoare","Unitate","Descriere"]
    for c, h in enumerate(headers_reguli, 1):
        ws_reguli.cell(2, c, h).font = hfont(10)
        ws_reguli.cell(2, c, h).fill = fill(C_MID)
        ws_reguli.cell(2, c, h).alignment = ctr()
        # Înlocuiește întregul bloc `reguli = [...]` și bucla de scriere cu:
        reguli = [
            # ── Secțiunea 1: Validare date orare / zi ──────────────────────
            ("─── VALIDARE DATE ───", "", "", ""),
            ("Minim ore / zi", MIN_ORE_ZI, "ore",
             "Zi validă = minim acest număr de ore înregistrate în ziua respectivă. "
             "Zilele sub acest prag sunt excluse din toate calculele (MZL, MZA)."),
            ("Minim zile / lună", MIN_ZILE_LUNA, "zile",
             "Lună validă pentru MZL = minim acest număr de zile valide (≥ MIN_ORE_ZI ore/zi)."),
            ("Alternativă 7 zile consecutive", MIN_ZILE_SAPT, "zile",
             f"Dacă luna nu are {MIN_ZILE_LUNA} zile valide, e considerată validă dacă există "
             f"cel puțin {MIN_ZILE_SAPT} zile consecutive valide. MZL se calculează pe acele 7 zile."),
            ("Date complete", "toate", "zile/lună",
             "Toate zilele lunii au date valide — MZL = suma/zile_în_lună."),
            ("Date parțiale", f"≥{MIN_ZILE_LUNA}", "zile/lună",
             f"Suficiente zile valide ({MIN_ZILE_LUNA}+) dar nu toate — MZL = suma/nr_zile_valide."),
            ("Date parțiale - 7 zile", "1", "săptămână",
             f"Fallback: cel puțin {MIN_ZILE_SAPT} zile consecutive valide — MZL = suma/7."),
            ("Nu există date (lună)", "0", "zile/lună",
             "Nicio zi validă în lună — luna exclusă complet din MZL și MZA."),
            ("Nu există date (zi)", "0", "ore/zi",
             "Ziua nu are nicio oră cu vehicule — marcată în Rezumat Zilnic cu roșu."),
            # ── Secțiunea 2: Date prelucrate (reconstrucție benzi) ──────────
            ("─── DATE PRELUCRATE ───", "", "", ""),
            ("Prag totalizator (T)", f">{int(PRAG_TOT_DP*100)}%", "din total bandă",
             f"Bandă = Totalizator dacă Clasa_15 (Others) depășește {int(PRAG_TOT_DP*100)}% "
             f"din totalul benzii în ziua respectivă. Clasele lipsesc → se reconstituie."),
            ("Bandă defectă (D)", "0", "vehicule",
             "Bandă = Defect dacă totalul este 0 vehicule în ziua respectivă."),
            ("Clasificator (C)", f"≤{int(PRAG_TOT_DP*100)}%", "Clasa_15",
             "Bandă normală — datele de clase sunt folosite direct, fără reconstituire."),
            ("Perechi benzi (2 benzi)", "B1↔B2", "sens opus",
             "Banda 1 și Banda 2 sunt sensuri opuse pe același amplasament. "
             "Banda D preia valorile perechii C."),
            ("Perechi benzi (4 benzi)", "B1↔B4, B2↔B3", "sens opus",
             "Sens 1: B1+B2; Sens 2: B3+B4. Perechile pentru reconstrucție: B1↔B4, B2↔B3."),
            ("Perechi benzi (6 benzi)", "B1↔B6, B2↔B5, B3↔B4", "sens opus",
             "Sens 1: B1+B2+B3; Sens 2: B4+B5+B6. Perechile pentru reconstrucție: B1↔B6, B2↔B5, B3↔B4."),
            ("Reconstituire T din C", "cls_n=(cls_n_pereche/tot_pereche)×tot_T", "clase",
             "Benzii T i se redistribuie clasele în proporțiile benzii pereche C, "
             "rotunjit la număr întreg."),
            ("Reconstituire D din C", "D=pereche", "toate clasele",
             "Banda D preia direct toate clasele benzii pereche C. "
             "Dacă ambele benzi sunt D → se caută altă bandă din același sens."),
            ("Zi neutilizabilă", "ambele D", "sens lipsă",
             "Dacă nu se poate reconstitui cel puțin un sens → ziua e marcată Neutilizabil "
             "și exclusă din MZL și MZA."),
            # ── Secțiunea 3: MZA ────────────────────────────────────────────
            ("─── MZA ───", "", "", ""),
            ("Minim luni valide / an", _MIN_LUNI_AN, "luni",
             f"MZA = medie aritmetică a MZL-urilor lunilor valide, dacă sunt cel puțin "
             f"{_MIN_LUNI_AN} luni valide din an. MZL = media zilnică a datelor prelucrate."),
            ("Fallback MZA — luna Mai", f"luna {_MIN_LUNI_AN_MAI}", "Mai",
             f"Dacă sunt sub {_MIN_LUNI_AN} luni valide, MZA = valoarea MZL a lunii Mai "
             f"(luna {_MIN_LUNI_AN_MAI}), dacă aceasta este validă."),
            ("Fallback MZA — luna Octombrie", f"luna {_MIN_LUNI_AN_OCT}", "Octombrie",
             f"Dacă nici luna Mai nu e validă, MZA = valoarea MZL a lunii Octombrie "
             f"(luna {_MIN_LUNI_AN_OCT}). Dacă nici Octombrie nu e validă → MZA lipsă."),
        ]
        # Scriem rândurile de reguli
        row_reguli = 3
        for reg, val, unit, desc in reguli:
            # Rând de secțiune (separator)
            if str(reg).startswith("───"):
                ws_reguli.merge_cells(f"A{row_reguli}:D{row_reguli}")
                _sc = ws_reguli.cell(row_reguli, 1, reg.replace("─── ", "").replace(" ───", ""))
                _sc.font      = Font(name='Arial', size=9, bold=True, color="FFFFFF")
                _sc.fill      = fill(C_MID)
                _sc.alignment = Alignment(horizontal='left', vertical='center')
                _sc.border    = brd
                ws_reguli.row_dimensions[row_reguli].height = 20
                row_reguli += 1
                continue
            rf = fill(C_LIGHT) if row_reguli % 2 == 0 else fill(C_WHITE)
            for c, v in enumerate([reg, val, unit, desc], 1):
                cell = ws_reguli.cell(row_reguli, c, v)
                cell.font = dfont(9)
                cell.fill = rf
                cell.border = brd
                cell.alignment = (ctr() if c <= 3
                                  else Alignment(horizontal='left',
                                                 vertical='center',
                                                 wrap_text=True))
            ws_reguli.row_dimensions[row_reguli].height = 30
            row_reguli += 1
        ws_reguli.column_dimensions['A'].width = 32
        ws_reguli.column_dimensions['B'].width = 22
        ws_reguli.column_dimensions['C'].width = 14
        ws_reguli.column_dimensions['D'].width = 75

        # ── Secțiunea „Fișiere sursă și perioadă procesată" ──────────────────
        # Calculăm rândul de start (după tabelul de reguli + 2 rânduri spațiu)
        start_sursa = len(reguli) + 3 + 2   # header la r=2, reguli incep la r=3

        # Titlu secțiune fișiere sursă
        ws_reguli.merge_cells(
            start_row=start_sursa, start_column=1,
            end_row=start_sursa, end_column=4)
        tc = ws_reguli.cell(start_sursa, 1,
                            "FIȘIERE SURSĂ  |  Perioadă procesată")
        tc.font      = Font(name='Arial', size=11, bold=True, color=C_WHITE)
        tc.fill      = fill(C_DARK)
        tc.alignment = ctr()
        ws_reguli.row_dimensions[start_sursa].height = 24

        # Header tabel fișiere sursă
        src_headers = ["Nr.", "Fișier sursă", "Cale completă", "Perioadă"]
        src_col_w   = [6, 42, 80, 28]
        for c_i, (h, w) in enumerate(zip(src_headers, src_col_w), 1):
            cell = ws_reguli.cell(start_sursa + 1, c_i, h)
            cell.font      = hfont(9)
            cell.fill      = fill(C_MID)
            cell.alignment = ctr()
            cell.border    = brd

        # Perioada corectă = perioada din Rezumat Zilnic (fallback global)
        perioada_globala = f"{first_date} — {last_date}"

        # Completăm rândurile cu fișierele sursă
        surse = source_files if source_files else []
        if not surse:
            surse_display = [("—", "—", perioada_globala)]
        else:
            surse_display = []
            for fp in surse:
                fname = os.path.basename(fp)
                fpath = os.path.abspath(fp)

                # ── Calculăm perioada per fișier din df ──────────────────────
                # source_file_periods este un dict opțional {basename → (min, max)}
                # transmis din parser; dacă nu există, folosim perioada globală
                perioad_fisier = perioada_globala
                if source_file_periods and fname in source_file_periods:
                    p_min, p_max = source_file_periods[fname]
                    perioad_fisier = f"{p_min} — {p_max}"

                surse_display.append((fname, fpath, perioad_fisier))

        for row_i, (fname, fpath, perioad) in enumerate(surse_display, 1):
            r = start_sursa + 1 + row_i
            rf = fill(C_LIGHT) if row_i % 2 == 1 else fill(C_WHITE)
            vals = [row_i, fname, fpath, perioad]
            aligns = [ctr(), Alignment(horizontal='left', vertical='center', wrap_text=False),
                      Alignment(horizontal='left', vertical='center', wrap_text=False),
                      ctr()]
            for c_i, (v, al) in enumerate(zip(vals, aligns), 1):
                cell = ws_reguli.cell(r, c_i, v)
                cell.font      = dfont(9)
                cell.fill      = rf
                cell.border    = brd
                cell.alignment = al
            ws_reguli.row_dimensions[r].height = 18

        # Lărgim coloana C pentru calea completă
        ws_reguli.column_dimensions['C'].width = 80

    wb.active = wb.sheetnames.index("Rezumat Zilnic")
    wb.save(excel_path)



# ══════════════════════════════════════════════════════════════════════════════
# PROCESARE COMPLETĂ UN FIȘIER
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# SCANARE RAPIDĂ — extrage site_id și nr. înregistrări fără procesare completă
# ══════════════════════════════════════════════════════════════════════════════
